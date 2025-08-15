import os
from datetime import datetime, timedelta
import calendar
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
import io
from collections import defaultdict
from datetime import datetime, timedelta
import logging
import logging
from datetime import datetime
import traceback
from dateutil import relativedelta
import calendar
import csv
import io
from datetime import datetime
from flask import Response
from jinja2 import Environment, FileSystemLoader
from functools import wraps
from flask_moment import Moment
from werkzeug.security import generate_password_hash, check_password_hash
from math import ceil
from flask import (
    Flask,
    flash,
    redirect,
    render_template,
    request,
    send_from_directory,
    session,
    url_for,
    jsonify, send_file, make_response
)
from flask_pymongo import PyMongo
from pymongo import MongoClient
from bson.objectid import ObjectId
import pymongo.errors
import uuid
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib import colors
from reportlab.lib.units import inch
from calendar import month_name
import json 
from collections import defaultdict

# Initialize Flask app
app = Flask(__name__)
logger = logging.getLogger(__name__)
moment = Moment(app)
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ===== Security Configurations =====
app.config['SECRET_KEY'] = '3179bc29135e94c099b732dcc43a0259c5c5fe2a981bbf5c17d78bd85768e735'
app.config['SESSION_COOKIE_SECURE'] = False  # Set to True in production with HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=1)

# ===== Database Configurations =====
app.config['MONGO_URI'] = "mongodb://localhost:27017/cost_reduction_tracker"
app.config['MONGO_CONNECT'] = False

# ===== File Upload Configurations =====
app.config['UPLOAD_FOLDER'] = 'static/exports'
app.config['ALLOWED_EXTENSIONS'] = {'csv', 'xlsx', 'xls'}
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app.config['UPLOAD_EXTENSIONS'] = ['.csv', '.xlsx', '.xls']

# ===== Performance Configurations =====
app.config['JSONIFY_PRETTYPRINT_REGULAR'] = False
app.config['TEMPLATES_AUTO_RELOAD'] = True

# Initialize MongoDB
mongo = PyMongo(app, connect=False)
db = mongo.db

# Initialize PyMongo
client = MongoClient(app.config['MONGO_URI'])
db = client.get_default_database()

# Context processor to make 'now' available in all templates
@app.context_processor
def inject_now():
    return {'now': datetime.now}

# Custom Jinja2 filter for datetime formatting
@app.template_filter("datetimeformat")
def datetimeformat(value, format="%Y-%m-%d %H:%M"):
    if isinstance(value, datetime):
        return value.strftime(format)
    return value

# Custom Jinja2 filter for rounding
@app.template_filter("round")
def round_filter(value, precision=0):
    try:
        return round(float(value), precision)
    except (ValueError, TypeError):
        return value

# Custom Jinja2 filter for number formatting with commas
@app.template_filter("numberformat")
def numberformat_filter(value, precision=2):
    try:
        num = float(value)
        if precision == 0:
            return f"{int(round(num, 0)):,}"
        return f"{round(num, precision):,.{precision}f}"
    except (ValueError, TypeError):
        return "N/A"
    

    
@app.route('/documentation')
def documentation():
    return render_template('documentation.html')  # or whatever template you want

@app.template_filter('to_formatted_currency')
def to_formatted_currency(value):
    try:
        return f"{float(value):,.2f}"
    except (ValueError, TypeError):
        return "N/A"
    
    
def login_required(role=None):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # Check if user is logged in
            if 'user_id' not in session:
                flash('Please log in to access this page', 'danger')
                return redirect(url_for('login'))
            
            # Check role if specified
            if role and session.get('role') != role:
                flash('You do not have permission to access this page', 'danger')
                # Redirect to appropriate dashboard based on user's role
                user_role = session.get('role')
                if user_role == 'admin':
                    return redirect(url_for('admin_dashboard'))
                elif user_role == 'manager':
                    return redirect(url_for('manager_dashboard'))
                else:
                    return redirect(url_for('user_dashboard'))
            
            # Call the original function with all arguments
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.route('/admin/currency')
@login_required(role='admin')
def admin_currency():
    current_year = datetime.now().year
    months = list(calendar.month_name)[1:]  # Get all month names (excluding empty first element)
    return render_template('admin/currency.html', 
                         current_year=current_year,
                         months=months)







# Helper Functions
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def read_excel_with_openpyxl(file_path):
    """Read Excel file using openpyxl and convert to DataFrame"""
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook.active
        data = []
        headers = []
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_num == 1:
                headers = [str(cell).strip() if cell is not None else f"column_{i}" for i, cell in enumerate(row, 1)]
            else:
                if any(cell is not None for cell in row):
                    padded_row = list(row) + [None] * (len(headers) - len(row))
                    data.append(padded_row[:len(headers)])
        workbook.close()
        df = pd.DataFrame(data, columns=headers)
        return df
    except Exception as e:
        raise Exception(f"Error reading Excel file: {str(e)}")

def create_styled_excel(data, filename):
    """Create a styled Excel file using openpyxl"""
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Factory Data"
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='800000', end_color='800000', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        if not data.empty:
            for col_num, header in enumerate(data.columns, 1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            for row_num, row_data in enumerate(data.itertuples(index=False), 2):
                for col_num, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_num, column=col_num, value=value)
                    cell.border = border
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        workbook.save(filename)
        workbook.close()
        return True
    except Exception as e:
        raise Exception(f"Error creating Excel file: {str(e)}")

def process_factory_data(file_path):
    """Process Excel/CSV file and split by Plant Number into separate files"""
    try:
        # Read the file
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = read_excel_with_openpyxl(file_path)
        
        # Clean column names and data
        df.columns = [str(col).strip() for col in df.columns]
        
        # Ensure Plant Number column exists
        if 'Plant_Number' not in df.columns:
            raise ValueError("'Plant Number' column is required in the uploaded file")
        
        # Define factory mappings
        factory_mappings = {
            6061: 'DPL1',
            6062: 'DPL2',
            6041: 'URIL'  # Add others as needed
        }
        
        results = []
        processed_files = []
        
        for plant_num, factory_code in factory_mappings.items():
            # Filter rows for this plant number
            factory_data = df[df['Plant_Number'] == plant_num].copy()
            
            if not factory_data.empty:
                # Generate filename (e.g., "DPL1_6061_data.xlsx")
                filename = f"{factory_code}_{plant_num}_data.xlsx"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                
                # Save to Excel
                create_styled_excel(factory_data, filepath)
                processed_files.append(filename)
                
                # Prepare data for MongoDB
                products_data = []
                for _, row in factory_data.iterrows():
                    product_dict = {
                        str(col): None if pd.isna(row[col]) else row[col] 
                        for col in factory_data.columns
                        if col != 'Plant_Number'  # Exclude Plant Number from products
                    }
                    products_data.append(product_dict)
                
                # Update MongoDB
                if db is not None:
                    db.factories.update_one(
                        {'plant_number': plant_num},  # Unique identifier
                        {
                            '$set': {
                                'factory_code': factory_code,
                                'plant_number': plant_num,
                                'products': products_data,
                                'upload_date': datetime.now(),
                                'file_path': filepath,
                                'record_count': len(factory_data)
                            }
                        },
                        upsert=True
                    )
                results.append(f"Processed {factory_code} with {len(factory_data)} records")
        
        return True, " | ".join(results), processed_files
    
    except Exception as e:
        return False, f"Processing error: {str(e)}", []

# Helper Functions
def get_currency():
    settings = db.settings.find_one({"name": "currency"})
    return settings["value"] if settings else "PKR"

def calculate_roi(project_data, months_to_show=60):
    project_type = project_data.get("project_type")
    machinery_cost = 0
    if project_type != "Kaizen":
        machinery_cost = float(project_data.get("machinery_cost", 0))
    params = project_data.get("project_parameters", [])
    monthly_saving_from_params = 0
    savings_by_param = {}
    
    for param_data in params:
        if isinstance(param_data, dict) and 'name' in param_data and 'before_value' in param_data and 'after_value' in param_data and 'cost' in param_data:
            try:
                before_value = float(param_data.get('before_value', 0))
                after_value = float(param_data.get('after_value', 0))
                cost = float(param_data.get('cost', 0))
                
                # New condition-based calculation
                if after_value > before_value:
                    # Improvement case (higher is better)
                    param_savings = (after_value - before_value) * cost
                elif after_value < before_value:
                    # Decline case (lower is worse)
                    param_savings = (before_value - after_value) * cost
                else:
                    # No change
                    param_savings = 0
                
                if param_savings > 0:
                    savings_by_param[param_data.get('name')] = param_savings
                monthly_saving_from_params += param_savings
            except (ValueError, TypeError):
                continue

    # Rest of the function remains the same...
    roi_months = []
    cumulative = 0
    roi_month = None
    months_to_show = int(months_to_show) if isinstance(months_to_show, (int, float)) and months_to_show > 0 else 60

    for i in range(1, months_to_show + 1):
        cumulative += monthly_saving_from_params
        is_roi_achieved_this_month = False
        if project_type != "Kaizen":
            if machinery_cost > 0 and roi_month is None and cumulative >= machinery_cost:
                is_roi_achieved_this_month = True
                roi_month = i
        elif project_type == "Kaizen" and monthly_saving_from_params > 0:
            is_roi_achieved_this_month = True if i == 1 else False
        
        roi_months.append({
            "month": f"Month {i}",
            "saving": monthly_saving_from_params,
            "cumulative": cumulative,
            "is_roi": is_roi_achieved_this_month
        })
        
        if project_type == "Kaizen" and monthly_saving_from_params > 0 and roi_month is None:
            roi_month = 1
            
    return {
        "savings": savings_by_param,
        "monthly_saving": monthly_saving_from_params,
        "months": roi_months,
        "roi_month": roi_month,
        "machinery_cost": machinery_cost,
        "is_kaizen": project_type == "Kaizen",
        "calculated_at": datetime.now().isoformat(),
        "parameters": params
    }
def export_roi_to_excel(project, roi_data, filename):
    """Export ROI calculation to Excel using openpyxl, including monthly tracking data."""
    try:
        workbook = Workbook()
        title_font = Font(bold=True, size=16, color='800000')
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='800000', end_color='800000', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        data_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        bold_style = Font(bold=True)
        summary_sheet = workbook.active
        summary_sheet.title = "Project Summary"
        summary_sheet['A1'] = "Project ROI Analysis"
        summary_sheet['A1'].font = title_font
        summary_sheet['A3'] = "Project Name:"
        summary_sheet['B3'] = project.get('project_name', 'N/A')
        summary_sheet['A4'] = "Factory:"
        summary_sheet['B4'] = project.get('factory_code', 'N/A')
        summary_sheet['A5'] = "Type:"
        summary_sheet['B5'] = project.get('project_type', 'N/A')
        summary_sheet['A6'] = "Created By:"
        summary_sheet['B6'] = project.get('created_by_username', 'N/A')
        summary_sheet['A7'] = "Created At:"
        created_at = project.get('created_at')
        if created_at and isinstance(created_at, datetime):
            summary_sheet['B7'] = created_at.strftime('%Y-%m-%d %H:%M')
        elif created_at and isinstance(created_at, str):
            summary_sheet['B7'] = datetime.fromisoformat(created_at).strftime('%Y-%m-%d %H:%M')
        else:
            summary_sheet['B7'] = 'N/A'
        summary_sheet['A9'] = "ROI Summary"
        summary_sheet['A9'].font = bold_style
        summary_sheet['A10'] = "Monthly Savings:"
        summary_sheet['B10'] = roi_data.get('monthly_saving', 0)
        summary_sheet['B10'].font = bold_style
        summary_sheet['A11'] = "Machinery Cost:"
        summary_sheet['B11'] = roi_data.get('machinery_cost', 0)
        summary_sheet['B11'].font = bold_style
        summary_sheet['A12'] = "ROI Month:"
        summary_sheet['B12'] = roi_data.get('roi_month', 'N/A')
        summary_sheet['B12'].font = bold_style
        roi_sheet = workbook.create_sheet(title="Monthly ROI")
        monthly_roi_headers = ['Month', 'Monthly Saving', 'Cumulative Saving']
        if not roi_data.get('is_kaizen'):
            monthly_roi_headers.append('ROI Achieved')
        for col, header in enumerate(monthly_roi_headers, 1):
            cell = roi_sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = data_border
        for row, month_data in enumerate(roi_data.get('months', []), 2):
            roi_sheet.cell(row=row, column=1, value=month_data['month']).border = data_border
            roi_sheet.cell(row=row, column=2, value=month_data['saving']).border = data_border
            roi_sheet.cell(row=row, column=3, value=month_data['cumulative']).border = data_border
            roi_sheet.cell(row=row, column=3).font = bold_style
            if not roi_data.get('is_kaizen'):
                roi_achieved_cell = roi_sheet.cell(row=row, column=4, value='Yes' if month_data['is_roi'] else 'No')
                roi_achieved_cell.border = data_border
                if month_data['is_roi']:
                    roi_achieved_cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        
        for sheet in workbook.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        filename = secure_filename(f"Project_ROI_{project.get('project_name', 'N/A')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        flash(f"Error exporting ROI to Excel: {e}", "danger")
        return redirect(url_for('user_project_details', project_id=project.get('project_id')))

# Helper Functions
def get_user_collection(role):
    if role == 'admin':
        return db.admins
    elif role == 'manager':
        return db.managers
    elif role == 'user':
        return db.users
    return None

def ensure_index(collection, field, **options):
    """Fixed ensure_index function to handle the database index creation properly"""
    try:
        # Check if index already exists
        existing_indexes = list(collection.list_indexes())
        index_name = f"{field}_1"  # MongoDB default naming convention
        
        # Check if index already exists
        for index in existing_indexes:
            if index.get('name') == index_name:
                print(f"Index on '{collection.name}.{field}' already exists.")
                return
        
        # Create the index if it doesn't exist
        collection.create_index(field, **options)
        print(f"Index on '{collection.name}.{field}' created successfully.")
        
    except pymongo.errors.OperationFailure as e:
        if "already exists" in str(e).lower():
            print(f"Index on '{collection.name}.{field}' already exists.")
        else:
            print(f"Error creating index on '{collection.name}.{field}': {e}")
    except Exception as e:
        print(f"Unexpected error creating index on '{collection.name}.{field}': {e}")

def initialize_database():
    """Initializes MongoDB collections and seeds default admin data."""
    if db is None:
        print("Database not connected, skipping initialization")
        return
    try:
        collections = ['users', 'admins', 'managers', 'factories', 'projects', 'reports', 'unit_costs', 'project_categories']
        existing_collections = db.list_collection_names()
        for col in collections:
            if col not in existing_collections:
                db.create_collection(col)
                print(f"Created collection: {col}")
        
        # Create indexes safely
        try:
            ensure_index(db.users, "email", unique=True, sparse=True)
            ensure_index(db.admins, "email", unique=True, sparse=True)
            ensure_index(db.managers, "email", unique=True, sparse=True)
            ensure_index(db.projects, "project_id", unique=True, sparse=True)
            ensure_index(db.project_categories, "name", unique=True, sparse=True)
            # Modified factory index to allow multiple nulls
            ensure_index(db.factories, "name", unique=True, sparse=True)
        except Exception as e:
            print(f"Index creation warning: {e}")
        
        if db.unit_costs.count_documents({}) == 0:
            default_unit_costs = {
                "DPL1": [
                    {"name": "Electricity", "unit": "KWH", "cost": 25, "operator": ">"},
                    {"name": "Gas", "unit": "MMBTU", "cost": 1500, "operator": ">"},
                    {"name": "Water", "unit": "Liter", "cost": 0.5, "operator": ">"}
                ],
                "DPL2": [
                    {"name": "Electricity", "unit": "KWH", "cost": 26, "operator": ">"},
                    {"name": "Gas", "unit": "MMBTU", "cost": 1550, "operator": ">"},
                    {"name": "Water", "unit": "Liter", "cost": 0.55, "operator": ">"}
                ],
                "URIL": [
                    {"name": "Electricity", "unit": "KWH", "cost": 24, "operator": ">"},
                    {"name": "Gas", "unit": "MMBTU", "cost": 1450, "operator": ">"},
                    {"name": "Water", "unit": "Liter", "cost": 0.48, "operator": ">"}
                ],
                "currency": "PKR"
            }
            db.unit_costs.insert_one(default_unit_costs)
            print("Default unit costs added.")
        
        default_categories = ["Kaizen", "MIPS", "Six Sigma", "Automation", "Infrastructure"]
        for cat_name in default_categories:
            db.project_categories.update_one(
                {"name": cat_name},
                {"$setOnInsert": {"name": cat_name, "created_at": datetime.now()}},
                upsert=True
            )
        print("Ensured default project categories exist.")
        
        # Create default admin user with correct credentials
        db.admins.update_one(
            {"email": "admin@example.com"},
            {"$set": {
                "username": "Admin User",
                "password": generate_password_hash("abcdef"),
                "email": "admin@example.com",
                "role": "admin",
                "status": 1,
                "created_at": datetime.now()
            }},
            upsert=True
        )
        print("Ensured default admin user is enabled.")
        print("Database initialization completed successfully!")
    except Exception as e:
        print(f"Database initialization error: {e}")

# Initialize the database when the app starts
with app.app_context():
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    initialize_database()

# Authentication Routes
@app.route("/")
def index():
    """Renders the welcome page."""
    return render_template("welcome.html")

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        try:
            # Get form data safely
            email_or_username = request.form.get('email', '').strip()
            password = request.form.get('password', '').strip()
            
            if not email_or_username or not password:
                flash('Please enter both email and password.', 'danger')
                return render_template('login.html')
            
            user = None
            user_role = None
            
            # Check all user collections by email or username
            # First, try to find by email
            admin = db.admins.find_one({'email': email_or_username})
            manager = db.managers.find_one({'email': email_or_username})
            regular_user = db.users.find_one({'email': email_or_username})
            
            if admin:
                user = admin
                user_role = 'admin'
            elif manager:
                user = manager
                user_role = 'manager'
            elif regular_user:
                user = regular_user
                user_role = 'user'
            
            # If not found by email, try to find by username
            if not user:
                admin = db.admins.find_one({'username': email_or_username})
                manager = db.managers.find_one({'username': email_or_username})
                regular_user = db.users.find_one({'username': email_or_username})

                if admin:
                    user = admin
                    user_role = 'admin'
                elif manager:
                    user = manager
                    user_role = 'manager'
                elif regular_user:
                    user = regular_user
                    user_role = 'user'

            if user and check_password_hash(user['password'], password):
                if user.get('status', 1) == 1:  # Default to active if status not set
                    # Set ALL required session variables
                    session.clear()  # Clear any existing session first
                    session['user_id'] = str(user['_id'])
                    session['username'] = user.get('username', user['email'])
                    session['email'] = user['email']  # Critical for change password
                    session['role'] = user_role
                    session.permanent = True  # Make session persistent
                    
                    flash('Logged in successfully!', 'success')
                    if user_role == 'admin':
                        return redirect(url_for('admin_dashboard'))
                    elif user_role == 'manager':
                        return redirect(url_for('manager_dashboard'))
                    else:
                        return redirect(url_for('user_dashboard'))
                else:
                    flash('Your account is disabled.', 'danger')
            else:
                flash('Invalid email or password.', 'danger')
        except Exception as e:
            flash('An error occurred during login. Please try again.', 'danger')
            print(f"Login error: {e}")
    
    return render_template('login.html')
@app.route('/manager/projects')
@login_required(role='manager')
def manager_projects():
    try:
        # Get filter parameters
        project_type_filter = request.args.get('project_type', '')
        year_filter = request.args.get('year', '')
        month_filter = request.args.get('month', '')
        
        # Build query based on filters
        query = {}
        if project_type_filter and project_type_filter != 'All Types':
            query['project_type'] = project_type_filter
            
        if year_filter and year_filter != 'All Years':
            try:
                year = int(year_filter)
                query['$expr'] = {
                    '$eq': [{'$year': '$created_at'}, year]
                }
            except ValueError:
                pass
                
        if month_filter and month_filter != 'All Months':
            try:
                month = int(month_filter)
                if '$expr' in query:
                    query['$expr']['$and'].append({'$eq': [{'$month': '$created_at'}, month]})
                else:
                    query['$expr'] = {
                        '$eq': [{'$month': '$created_at'}, month]
                    }
            except ValueError:
                pass

        # Get filtered projects
        projects = list(db.projects.find(query).sort('created_at', -1))
        
        # Calculate summary statistics
        total_projects = len(projects)
        total_estimated_savings = sum(p.get('estimated_savings', 0) for p in projects)
        total_machinery_cost = sum(p.get('machinery_cost', 0) for p in projects)
        approved_projects = len([p for p in projects if p.get('status') == 'Approved'])
        pending_projects = len([p for p in projects if p.get('status') == 'Pending'])
        
        # Get project categories for filter dropdown
        project_categories = list(db.project_categories.find().sort('name', 1))
        
        # Get currency
        unit_costs_doc = db.unit_costs.find_one()
        currency = unit_costs_doc.get('currency', 'PKR') if unit_costs_doc else 'PKR'

        return render_template('manager/projects.html',
                            projects=projects,
                            project_categories=project_categories,
                            currency=currency,
                            current_year=datetime.now().year,
                            month_names=list(calendar.month_name)[1:],
                            now=datetime.now(),
                            selected_project_type=project_type_filter,
                            selected_year=year_filter,
                            selected_month=month_filter,
                            summary_stats={
                                'total_projects': total_projects,
                                'total_estimated_savings': total_estimated_savings,
                                'total_machinery_cost': total_machinery_cost,
                                'approved_projects': approved_projects,
                                'pending_projects': pending_projects
                            })
    except Exception as e:
        flash(f'Error loading projects: {str(e)}', 'danger')
        return redirect(url_for('manager_dashboard'))

@app.route('/export/projects', methods=['GET'])
@login_required(role='manager')
def export_projects():
    try:
        # Get filter parameters from request
        project_type_filter = request.args.get('project_type', '')
        year_filter = request.args.get('year', '')
        month_filter = request.args.get('month', '')
        
        # Build query based on filters
        query = {}
        if project_type_filter and project_type_filter != 'All Types':
            query['project_type'] = project_type_filter
            
        if year_filter and year_filter != 'All Years':
            try:
                year = int(year_filter)
                query['$expr'] = {
                    '$eq': [{'$year': '$created_at'}, year]
                }
            except ValueError:
                pass
                
        if month_filter and month_filter != 'All Months':
            try:
                month = int(month_filter)
                if '$expr' in query:
                    query['$expr']['$and'].append({'$eq': [{'$month': '$created_at'}, month]})
                else:
                    query['$expr'] = {
                        '$eq': [{'$month': '$created_at'}, month]
                    }
            except ValueError:
                pass

        # Fetch filtered projects
        projects = list(db.projects.find(query).sort('created_at', -1))
        
        if not projects:
            flash('No projects found to export with current filters.', 'warning')
            return redirect(url_for('manager_projects'))

        # Calculate summary statistics
        total_estimated_savings = sum(p.get('estimated_savings', 0) for p in projects)
        total_machinery_cost = sum(p.get('machinery_cost', 0) for p in projects)
        approved_projects = len([p for p in projects if p.get('status') == 'Approved'])
        pending_projects = len([p for p in projects if p.get('status') == 'Pending'])

        # Create Excel workbook
        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Projects"
        
        # Headers
        headers = [
            "Project ID", "Project Name", "Type", "Factory", 
            "Status", "Created By", "Created At", "Start Date", 
            "End Date", "Estimated Savings", "Machinery Cost"
        ]
        
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)
            sheet.cell(row=1, column=col_num).font = Font(bold=True)
        
        # Data
        for row_num, project in enumerate(projects, 2):
            created_at = project.get('created_at')
            if isinstance(created_at, str):
                try:
                    created_at = datetime.fromisoformat(created_at)
                except ValueError:
                    created_at = None
            
            sheet.cell(row=row_num, column=1, value=project.get('project_id', 'N/A'))
            sheet.cell(row=row_num, column=2, value=project.get('project_name', 'N/A'))
            sheet.cell(row=row_num, column=3, value=project.get('project_type', 'N/A'))
            sheet.cell(row=row_num, column=4, value=project.get('factory_code', 'N/A'))
            sheet.cell(row=row_num, column=5, value=project.get('status', 'N/A'))
            sheet.cell(row=row_num, column=6, value=project.get('created_by_username', 'N/A'))
            sheet.cell(row=row_num, column=7, value=created_at.strftime('%Y-%m-%d %H:%M') if created_at else 'N/A')
            sheet.cell(row=row_num, column=8, value=project.get('start_date', 'N/A'))
            sheet.cell(row=row_num, column=9, value=project.get('end_date', 'N/A'))
            sheet.cell(row=row_num, column=10, value=project.get('estimated_savings', 0))
            sheet.cell(row=row_num, column=11, value=project.get('machinery_cost', 0))
        
        # Add summary section
        summary_row = row_num + 2
        sheet.cell(row=summary_row, column=1, value="SUMMARY STATISTICS").font = Font(bold=True)
        sheet.cell(row=summary_row+1, column=1, value="Total Projects")
        sheet.cell(row=summary_row+1, column=2, value=len(projects))
        sheet.cell(row=summary_row+2, column=1, value="Total Estimated Savings")
        sheet.cell(row=summary_row+2, column=2, value=total_estimated_savings)
        sheet.cell(row=summary_row+3, column=1, value="Total Machinery Cost")
        sheet.cell(row=summary_row+3, column=2, value=total_machinery_cost)
        sheet.cell(row=summary_row+4, column=1, value="Approved Projects")
        sheet.cell(row=summary_row+4, column=2, value=approved_projects)
        sheet.cell(row=summary_row+5, column=1, value="Pending Projects")
        sheet.cell(row=summary_row+5, column=2, value=pending_projects)
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(output)
        output.seek(0)
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=projects_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response
    
    except Exception as e:
        flash(f'Error exporting projects: {str(e)}', 'danger')
        return redirect(url_for('manager_projects'))

@app.route('/logout')
def logout():
    session.clear()  # Completely wipe the session
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/change-password', methods=['GET', 'POST'])
def change_password():
    if 'user_id' not in session:
        flash('Please log in to access this page.', 'danger')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        try:
            current_password = request.form.get('current_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')

            if not all([current_password, new_password, confirm_password]):
                flash('All fields are required.', 'danger')
                return redirect(url_for('change_password'))

            if new_password != confirm_password:
                flash('New passwords do not match.', 'danger')
                return redirect(url_for('change_password'))

            if len(new_password) < 6:
                flash('Password must be at least 6 characters long.', 'danger')
                return redirect(url_for('change_password'))

            user_collection = get_user_collection(session['role'])
            user = user_collection.find_one({'_id': ObjectId(session['user_id'])})
            
            if not user or not check_password_hash(user['password'], current_password):
                flash('Current password is incorrect.', 'danger')
                return redirect(url_for('change_password'))

            user_collection.update_one(
                {'_id': ObjectId(session['user_id'])},
                {'$set': {'password': generate_password_hash(new_password)}}
            )
            
            flash('Password updated successfully!', 'success')
            
            # Redirect based on role
            if session['role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            elif session['role'] == 'manager':
                return redirect(url_for('manager_dashboard'))
            else:
                return redirect(url_for('user_dashboard'))

        except Exception as e:
            flash('An error occurred. Please try again.', 'danger')
            return redirect(url_for('change_password'))

    return render_template('change_password.html')

# User Routes
@app.route('/user/dashboard')
@login_required(role='user')
def user_dashboard():
    user_id = session['user_id']
    projects = list(db.projects.find({'created_by_user_id': user_id}).sort('created_at', -1))

    # Basic counts
    total_projects = len(projects)
    pending_projects = len([p for p in projects if p.get('status') == 'Pending'])
    approved_projects = len([p for p in projects if p.get('status') == 'Approved'])
    rejected_projects = len([p for p in projects if p.get('status') == 'Rejected'])
    total_savings = sum(p.get('estimated_savings', 0) for p in projects if p.get('status') == 'Approved')

    # Status distribution for pie chart (ensure we always have 3 values)
    status_distribution = [
        approved_projects or 0,
        pending_projects or 0,
        rejected_projects or 0
    ]

    # ROI Analysis data (ensure we always have 3 values)
    roi_distribution = [
        approved_projects or 0,
        int((total_projects or 1) * 0.35),
        (total_projects or 1) - (approved_projects or 0) - int((total_projects or 1) * 0.35)
    ]

    # Monthly performance data (ensure we always have data)
    monthly_labels = []
    monthly_projects = []
    monthly_savings = []

    for i in range(5, -1, -1):
        month = datetime.now() - timedelta(days=30 * i)
        month_str = month.strftime('%b')
        monthly_labels.append(month_str)

        month_projects_count = len([
            p for p in projects
            if p.get('created_at') and
            isinstance(p['created_at'], datetime) and
            p['created_at'].month == month.month and
            p['created_at'].year == month.year
        ])
        monthly_projects.append(month_projects_count or 0)

        month_savings_total = sum(
            p.get('estimated_savings', 0)
            for p in projects
            if p.get('status') == 'Approved' and
            p.get('created_at') and
            isinstance(p['created_at'], datetime) and
            p['created_at'].month == month.month and
            p['created_at'].year == month.year
        )
        monthly_savings.append((month_savings_total or 0) // 1000)

    # Fallback data if no projects exist
    if not any(monthly_projects):
        monthly_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']
        monthly_projects = [2, 3, 5, 4, 6, 7]
        monthly_savings = [12, 15, 18, 14, 20, 25]

    # Savings breakdown (sample data)
    savings_categories = ['Materials', 'Labor', 'Energy', 'Logistics', 'Overhead']
    savings_values = [45, 28, 15, 22, 18]

    return render_template('user/user_dashboard.html',
                         projects=projects,
                         total_projects=total_projects,
                         pending_projects=pending_projects,
                         approved_projects=approved_projects,
                         rejected_projects=rejected_projects,
                         total_savings=total_savings,
                         status_distribution=status_distribution,
                         roi_distribution=roi_distribution,
                         monthly_labels=monthly_labels,
                         monthly_projects=monthly_projects,
                         monthly_savings=monthly_savings,
                         savings_categories=savings_categories,
                         savings_values=savings_values)
    
@app.route('/user/project_type', methods=['GET', 'POST'])
@login_required(role='user')
def project_type():
    project_categories = list(db.project_categories.find().sort('name', 1))
    project_data = session.get('current_project_data', {})
    if request.method == 'POST':
        project_type = request.form['project_type']
        session['current_project_data'] = {'project_type': project_type}
        return redirect(url_for('project_details_input'))
    return render_template('user/project_type.html', categories=project_categories, project_data=project_data)

@app.route('/user/project_details_input', methods=['GET', 'POST'])
@login_required(role='user')
def project_details_input():
    project_data = session.get('current_project_data', {})
    
    if request.method == 'POST':
        try:
            # Validate required fields
            required_fields = ['project_name', 'project_description', 'start_date', 'end_date']
            for field in required_fields:
                if field not in request.form or not request.form[field].strip():
                    flash(f"Missing or invalid {field.replace('_', ' ')}", 'danger')
                    return redirect(request.url)
            
            # Update project data
            project_data.update({
                'project_name': request.form['project_name'].strip(),
                'project_description': request.form['project_description'].strip(),
                'start_date': request.form['start_date'],
                'end_date': request.form['end_date']
            })
            
            # Validate dates
            if project_data['start_date'] > project_data['end_date']:
                flash("End date must be after start date", 'danger')
                return redirect(request.url)
                
            session['current_project_data'] = project_data
            return redirect(url_for('factory_select'))
            
        except Exception as e:
            flash(f"An error occurred: {str(e)}", 'danger')
            return redirect(request.url)
    
    # GET request - render template
    return render_template('user/project_details_input.html', project_data=project_data)

@app.route('/user/factory_select', methods=['GET', 'POST'])
@login_required(role='user')
def factory_select():
    project_data = session.get('current_project_data', {})
    factories_from_db = list(db.factories.find({}, {'factory_code': 1, 'plant_number': 1}))
    if not factories_from_db:
        flash("No factory data uploaded yet. Please contact an administrator to upload factory data.", 'warning')
        display_factories = []
    else:
        display_factories = []
        for f in factories_from_db:
            display_factories.append({'code': f['factory_code'], 'name': f['factory_code'] + ' (Plant ' + str(f['plant_number']) + ')'})
    if request.method == 'POST':
        factory_code = request.form['factory_code']
        project_data['factory_code'] = factory_code
        session['current_project_data'] = project_data
        return redirect(url_for('product_select'))
    return render_template('user/factory_select.html', factories=display_factories, project_data=project_data)

@app.route('/user/product_select', methods=['GET', 'POST'])
@login_required(role='user')
def product_select():
    project_data = session.get('current_project_data', {})
    factory_code = project_data.get('factory_code')
    
    if not factory_code:
        flash("Please select a factory first.", 'warning')
        return redirect(url_for('factory_select'))

    factory_doc = db.factories.find_one({'factory_code': factory_code})
    products = []
    
    if factory_doc:
        if 'products' in factory_doc:
            unique_products = {}
            for p_data in factory_doc['products']:
                # Convert all keys to lowercase for case-insensitive matching
                p_data_lower = {k.lower(): v for k, v in p_data.items()}
                
                # Find product name using possible key variations
                product_name = None
                for key in ['product_name', 'product', 'name', 'item', 'description']:
                    if key in p_data_lower and p_data_lower[key]:
                        product_name = str(p_data_lower[key]).strip()
                        if product_name:  # Only break if we found a non-empty name
                            break
                
                if product_name:
                    unique_products[product_name] = {
                        'product_id': product_name,  # Using name as ID for simplicity
                        'name': product_name
                    }
            
            products = list(unique_products.values())
    
    if not products:
        flash_message = (f"No products found for factory {factory_code}. "
                        f"Please ensure: (1) Factory data is uploaded via Admin panel, "
                        f"(2) Contains a column like 'Product_Name', 'Product', or 'Name', "
                        f"(3) The column has actual values")
        flash(flash_message, 'info')
    
    if request.method == 'POST':
        selected_products = request.form.getlist('selected_products')
        if not selected_products:
            flash('Please select at least one product.', 'danger')
        else:
            project_data['selected_products'] = selected_products
            session['current_project_data'] = project_data
            return redirect(url_for('model_select'))
    
    return render_template('user/product_select.html', 
                     products=products,
                     project_data=project_data)

@app.route('/user/model_select', methods=['GET', 'POST'])
@login_required(role='user')
def model_select():
    project_data = session.get('current_project_data', {})
    selected_products = project_data.get('selected_products', [])
    factory_code = project_data.get('factory_code')
    if not factory_code or not selected_products:
        flash("Please complete previous steps (factory and product selection).", 'warning')
        return redirect(url_for('product_select'))
    factory_doc = db.factories.find_one({'factory_code': factory_code})
    product_models = {}
    if factory_doc and 'products' in factory_doc:
        for selected_prod_name in selected_products:
            unique_models_for_product = {}
            for p_data in factory_doc['products']:
                p_data_lower_keys = {k.lower(): v for k, v in p_data.items()}
                current_product_name = None
                for key_candidate in ['product_name', 'product', 'name', 'item', 'description']:
                    if key_candidate in p_data_lower_keys and p_data_lower_keys[key_candidate] is not None:
                        product_name_candidate = str(p_data_lower_keys[key_candidate]).strip()
                        if product_name_candidate:
                            current_product_name = product_name_candidate
                            break
                if current_product_name and current_product_name == selected_prod_name:
                    model_name = None
                    for key_candidate in ['model', 'variant', 'sku', 'id', 'name']:
                        if key_candidate in p_data_lower_keys and p_data_lower_keys[key_candidate] is not None:
                            model_name_candidate = str(p_data_lower_keys[key_candidate]).strip()
                            if model_name_candidate:
                                model_name = model_name_candidate
                                break
                    if not model_name:
                        model_name = f"Generic Model for {selected_prod_name}"
                    if model_name not in unique_models_for_product:
                        unique_models_for_product[model_name] = {
                            'model_id': model_name,
                            'name': model_name,
                            'description': p_data_lower_keys.get('description', 'No description available')
                        }
            models_list = list(unique_models_for_product.values())
            if not models_list:
                models_list.append({
                    'model_id': f"GenericModel_{selected_prod_name}",
                    'name': f"Generic Model for {selected_prod_name}",
                    'description': 'No specific model data found in uploaded file for this product.'
                })
            product_models[selected_prod_name] = models_list
    if request.method == 'POST':
        selected_models = request.form.getlist('selected_models')
        if not selected_models:
            flash('Please select at least one model.', 'danger')
            return render_template('user/model_select.html', product_models=product_models, project_data=project_data)
        project_data['selected_models'] = selected_models
        session['current_project_data'] = project_data
        return redirect(url_for('project_parameters'))
    return render_template('user/model_select.html', product_models=product_models, project_data=project_data)



@app.route('/user/project_parameters', methods=['GET', 'POST'])
@login_required(role='user')
def project_parameters():
    try:
        # Initialize project data from session or new project
        project_data = session.get('current_project_data', {})
        project_id_from_url = request.args.get('project_id')
        
        # Get factory code - handle missing case
        factory_code = project_data.get('factory_code')
        if not factory_code and project_id_from_url:
            # Try to get from database if editing existing project
            existing_project = db.projects.find_one({'project_id': project_id_from_url})
            if existing_project:
                factory_code = existing_project.get('factory_code')
                project_data['factory_code'] = factory_code
                session['current_project_data'] = project_data
        
        if not factory_code:
            flash("Factory not specified in project data", 'danger')
            return redirect(url_for('user_dashboard'))

        # Fetch unit costs for the specific factory
        unit_costs_doc = db.unit_costs.find_one({})
        if not unit_costs_doc:
            flash("No unit costs configured in system", 'danger')
            return redirect(url_for('user_dashboard'))
        
        # Get parameters for the specific factory (like DPL1, DPL2, etc.)
        factory_params = unit_costs_doc.get(factory_code, [])
        currency = unit_costs_doc.get('currency', 'PKR')

        # If editing existing project
        if project_id_from_url:
            existing_project = db.projects.find_one({
                'project_id': project_id_from_url, 
                'created_by_user_id': session['user_id']
            })
            
            if not existing_project:
                flash("Project not found or you don't have permission to edit.", 'danger')
                return redirect(url_for('user_dashboard'))
                
            # Update project_data with existing project details
            project_data.update({
                'project_type': existing_project.get('project_type'),
                'project_name': existing_project.get('project_name'),
                'factory_code': existing_project.get('factory_code'),
                'project_parameters': existing_project.get('project_parameters', []),
                'machinery_cost': existing_project.get('machinery_cost', 0),
                'project_id': existing_project.get('project_id'),
                'status': existing_project.get('status', 'Pending')
            })
            session['current_project_data'] = project_data
            flash(f"Editing project: {existing_project.get('project_name')}", 'info')

        # Prepare display parameters
        display_params = []
        for param in project_data.get('project_parameters', []):
            param_name = param.get('name', '')
            matching_param = next(
                (p for p in factory_params if p['name'].lower() == param_name.lower()), 
                None
            )
            
            display_params.append({
                'name': param_name,
                'unit': param.get('unit', matching_param['unit'] if matching_param else ''),
                'cost': param.get('cost', matching_param['cost'] if matching_param else 0),
                'operator': param.get('operator', matching_param['operator'] if matching_param else '>'),
                'before_value': param.get('before_value', 0),
                'after_value': param.get('after_value', 0),
                'is_custom': not bool(matching_param)
            })

        if request.method == 'POST':
            # Process form data
            param_names = request.form.getlist('param_name[]')
            param_units = request.form.getlist('param_unit[]')
            param_costs = request.form.getlist('param_cost[]')
            param_operators = request.form.getlist('param_operator[]')
            param_before_values = request.form.getlist('param_before_value[]')
            param_after_values = request.form.getlist('param_after_value[]')
            
            # Validate input lengths
            if len({len(param_names), len(param_units), len(param_costs), 
                   len(param_operators), len(param_before_values), len(param_after_values)}) != 1:
                flash("Mismatched parameter data. Please check all fields.", 'danger')
                return redirect(url_for('project_parameters', project_id=project_id_from_url))
                
            # Process parameters
            parameters = []
            for i in range(len(param_names)):
                try:
                    parameters.append({
                        'name': param_names[i].strip(),
                        'unit': param_units[i].strip(),
                        'cost': float(param_costs[i]),
                        'operator': param_operators[i].strip(),
                        'before_value': float(param_before_values[i]),
                        'after_value': float(param_after_values[i]),
                        'is_custom': request.form.get(f'is_custom_param_{i}', 'false') == 'true'
                    })
                except ValueError as e:
                    flash(f"Invalid number in parameter {i+1}: {str(e)}", 'danger')
                    return redirect(url_for('project_parameters', project_id=project_id_from_url))
            
            # Process machinery cost
            machinery_cost = 0
            if project_data.get('project_type') != "Kaizen":
                try:
                    machinery_cost = float(request.form.get('machinery_cost', '0'))
                except ValueError:
                    flash("Invalid machinery cost value", 'danger')
                    return redirect(url_for('project_parameters', project_id=project_id_from_url))
            
            # Update project data
            project_data.update({
                'project_parameters': parameters,
                'machinery_cost': machinery_cost
            })
            session['current_project_data'] = project_data
            
            # Update in DB if editing existing project
            if project_id_from_url:
                db.projects.update_one(
                    {'project_id': project_id_from_url},
                    {'$set': {
                          'project_parameters': parameters,
                        'machinery_cost': machinery_cost,
                        'last_modified': datetime.utcnow()
                    }}
                )
                flash("Project parameters updated successfully", 'success')
            
            return redirect(url_for('project_timeline_input'))
        
        return render_template('user/project_parameters.html',
            project_type=project_data.get('project_type'),
            project_data=project_data,
            display_params=display_params,
            machinery_cost_value=project_data.get('machinery_cost', 0),
            currency=currency,
            all_unit_costs=factory_params  # This now contains only the factory-specific parameters
        )
        
    except Exception as e:
        print(f"ERROR in project_parameters: {str(e)}")
        print(traceback.format_exc())
        flash("An error occurred while processing project parameters", 'danger')
        return redirect(url_for('user_dashboard'))
    
    
    
@app.route('/user/project_timeline_input', methods=['GET', 'POST'])
@login_required(role='user')
def project_timeline_input():
    project_data = session.get('current_project_data', {})
    if not project_data:
        flash("Please start a new project or continue an existing one.", "warning")
        return redirect(url_for('user_dashboard'))

    # Initialize monthly data if missing
    if 'monthlyData' not in project_data or not project_data['monthlyData']:
        start_date_str = project_data.get('start_date')
        end_date_str = project_data.get('end_date')
        
        if start_date_str and end_date_str:
            try:
                start_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date_str, '%Y-%m-%d')
                monthly_data = []
                current_month_dt = start_dt.replace(day=1)
                
                # Calculate default monthly saving from project parameters
                estimated_savings_per_month = 0
                for param in project_data.get('project_parameters', []):
                    try:
                        diff = param['before_value'] - param['after_value'] if param['operator'] == '>' else param['after_value'] - param['before_value']
                        estimated_savings_per_month += diff * param['cost']
                    except (KeyError, TypeError) as e:
                        flash(f"Invalid parameter data: {str(e)}", "danger")
                        return redirect(url_for('project_parameters'))

                while current_month_dt <= end_dt:
                    try:
                        month_name = current_month_dt.strftime('%B %Y')
                        days_in_month = calendar.monthrange(current_month_dt.year, current_month_dt.month)[1]
                        
                        monthly_data.append({
                            'month': month_name,
                            'year': current_month_dt.year,
                            'monthNumber': current_month_dt.month,
                            'days': [{'day': day, 'value': 0.0, 'notes': ''} for day in range(1, days_in_month + 1)],
                            'totalValue': estimated_savings_per_month,
                            'userEnteredMonthly': True
                        })
                    except Exception as e:
                        flash(f"Error initializing month data: {str(e)}", "danger")
                        return redirect(url_for('project_parameters'))
                    
                    # Move to next month
                    if current_month_dt.month == 12:
                        current_month_dt = current_month_dt.replace(year=current_month_dt.year + 1, month=1)
                    else:
                        current_month_dt = current_month_dt.replace(month=current_month_dt.month + 1)
                
                project_data['monthlyData'] = monthly_data
                session['current_project_data'] = project_data
            except Exception as e:
                flash(f"Error initializing timeline: {str(e)}", "danger")
                return redirect(url_for('project_parameters'))

    # Calculate max value for visualization
    max_value = max((month['totalValue'] for month in project_data.get('monthlyData', [])), default=0)

    if request.method == 'POST':
        try:
            monthly_data_from_form = []
            month_names = request.form.getlist('month_name[]')
            month_years = request.form.getlist('month_year[]')
            month_numbers = request.form.getlist('month_number[]')
            
            # Process each month
            for i in range(len(month_names)):
                month_name = month_names[i]
                month_year = int(month_years[i])
                month_number = int(month_numbers[i])
                is_monthly_entry = request.form.get(f'user_entered_monthly[{i}]') == 'true'
                
                if is_monthly_entry:
                    # Monthly total entry
                    monthly_total = float(request.form.get(f'monthly_total[{i}]', 0))
                    monthly_data_from_form.append({
                        'month': month_name,
                        'year': month_year,
                        'monthNumber': month_number,
                        'days': [],
                        'totalValue': monthly_total,
                        'userEnteredMonthly': True
                    })
                else:
                    # Daily breakdown
                    days_data = []
                    total_value = 0
                    days_in_month = calendar.monthrange(month_year, month_number)[1]
                    
                    for day in range(1, days_in_month + 1):
                        day_value = float(request.form.get(f'day_value_{i}_{day}', 0))
                        day_notes = request.form.get(f'day_notes_{i}_{day}', '').strip()
                        days_data.append({
                            'day': day,
                            'value': day_value,
                            'notes': day_notes
                        })
                        total_value += day_value
                    
                    monthly_data_from_form.append({
                        'month': month_name,
                        'year': month_year,
                        'monthNumber': month_number,
                        'days': days_data,
                        'totalValue': total_value,
                        'userEnteredMonthly': False
                    })

            # Update project data in session
            project_data['monthlyData'] = monthly_data_from_form
            
            # Calculate financial metrics
            estimated_savings = sum(month['totalValue'] for month in monthly_data_from_form)
            machinery_cost = project_data.get('machinery_cost', 0)
            
            # ROI calculation (annualized)
            roi = (estimated_savings * 12 / machinery_cost) * 100 if machinery_cost > 0 else 0
            
            # Payback period (in months)
            payback_period = machinery_cost / estimated_savings if estimated_savings > 0 else 0
            
            project_data.update({
                'estimated_savings': estimated_savings,
                'roi': round(roi, 2),
                'payback_period': round(payback_period, 2),
                'roi_data': {
                    'monthly_saving': estimated_savings,
                    'roi_month': round(payback_period, 2),
                    'calculated_at': datetime.now().isoformat(),
                    'machinery_cost': machinery_cost,
                    'is_kaizen': project_data.get('project_type') == 'Kaizen',
                    'parameters': project_data.get('project_parameters', [])
                }
            })

            # Save to database (MongoDB)
            if 'project_id' not in project_data:
                project_id = str(uuid.uuid4())
                project_data.update({
                    'project_id': project_id,
                    'status': 'Pending',
                    'created_by_user_id': session['user_id'],
                    'created_by_username': session['username'],
                    'created_at': datetime.now(),
                    'updated_at': datetime.now(),
                    'actual_savings': 0.0,
                    'actualMonthlyData': [{
                        'month': month['month'],
                        'year': month['year'],
                        'monthNumber': month['monthNumber'],
                        'days': [{'day': day['day'], 'value': 0.0, 'notes': ''} for day in month.get('days', [])],
                        'totalValue': 0.0,
                        'userEnteredMonthly': month['userEnteredMonthly']
                    } for month in monthly_data_from_form]
                })
                db.projects.insert_one(project_data)
                flash('Project created successfully!', 'success')
            else:
                db.projects.update_one(
                    {'project_id': project_data['project_id']},
                    {'$set': {
                        'monthlyData': monthly_data_from_form,
                        'estimated_savings': estimated_savings,
                        'roi': round(roi, 2),
                        'payback_period': round(payback_period, 2),
                        'roi_data': project_data['roi_data'],
                        'updated_at': datetime.now()
                    }}
                )
                flash('Project updated successfully!', 'success')

            session.pop('current_project_data', None)
            return redirect(url_for('user_project_details', project_id=project_data['project_id']))
        
        except Exception as e:
            flash(f"Error processing form: {str(e)}", "danger")
            return redirect(url_for('project_timeline_input'))

    # Get currency for display
    try:
        unit_costs_doc = db.unit_costs.find_one()
        currency = unit_costs_doc.get('currency', 'PKR') if unit_costs_doc else 'PKR'
    except Exception as e:
        flash(f"Error loading currency: {str(e)}", "warning")
        currency = 'PKR'
    
    return render_template('user/project_timeline_input.html',
                         project_data=project_data,
                         currency=currency,
                         max_value=max_value)

@app.route('/user/actual_timeline_input/<project_id>', methods=['GET', 'POST'])
@login_required(role='user')
def actual_timeline_input(project_id):
    project = db.projects.find_one({'project_id': project_id, 'created_by_user_id': session['user_id']})
    if not project:
        flash('Project not found or you do not have access.', 'danger')
        return redirect(url_for('user_dashboard'))
    
    if project.get('status') != 'Approved':
        flash('You can only enter actual values for approved projects.', 'warning')
        return redirect(url_for('user_project_details', project_id=project_id))
    
    # Initialize actual data if missing
    monthly_data_for_actuals = project.get('actualMonthlyData', [])
    if not monthly_data_for_actuals:
        monthly_data_for_actuals = []
        for month_data in project.get('monthlyData', []):
            monthly_data_for_actuals.append({
                'month': month_data['month'],
                'year': month_data['year'],
                'monthNumber': month_data['monthNumber'],
                'days': [{'day': day['day'], 'value': 0.0, 'notes': ''} for day in month_data.get('days', [])],
                'totalValue': 0.0,
                'userEnteredMonthly': month_data.get('userEnteredMonthly', True)
            })

    # Calculate max value for visualization
    max_actual_value = max((month['totalValue'] for month in monthly_data_for_actuals), default=0)

    if request.method == 'POST':
        try:
            actual_monthly_data = []
            month_names = request.form.getlist('month_name[]')
            month_years = request.form.getlist('month_year[]')
            month_numbers = request.form.getlist('month_number[]')
            
            for i in range(len(month_names)):
                month_entry = {
                    'month': month_names[i],
                    'year': int(month_years[i]),
                    'monthNumber': int(month_numbers[i]),
                    'userEnteredMonthly': request.form.get(f'user_entered_monthly[{i}]') == 'true',
                    'days': []
                }
                
                if month_entry['userEnteredMonthly']:
                    # Monthly total entry
                    month_entry['totalValue'] = float(request.form.get(f'total_value[{i}]', 0))
                else:
                    # Daily breakdown
                    month_entry['totalValue'] = 0
                    days_in_month = calendar.monthrange(month_entry['year'], month_entry['monthNumber'])[1]
                    for day_num in range(1, days_in_month + 1):
                        day_value = float(request.form.get(f'day_value_{i}_{day_num}', 0))
                        day_notes = request.form.get(f'day_notes_{i}_{day_num}', '').strip()
                        month_entry['days'].append({
                            'day': day_num,
                            'value': day_value,
                            'notes': day_notes
                        })
                        month_entry['totalValue'] += day_value
                
                actual_monthly_data.append(month_entry)
            
            # Update project (MongoDB)
            db.projects.update_one(
                {'_id': project['_id']},
                {'$set': {
                    'actualMonthlyData': actual_monthly_data,
                    'updated_at': datetime.now(),
                    'actual_savings': sum(m['totalValue'] for m in actual_monthly_data)
                }}
            )
            flash('Actual data saved successfully!', 'success')
            return redirect(url_for('milestone_view', project_id=project_id))
            
        except Exception as e:
            flash(f"Error saving data: {str(e)}", 'danger')
    
    # Get currency
    unit_costs_doc = db.unit_costs.find_one()
    currency = unit_costs_doc.get('currency', 'PKR') if unit_costs_doc else 'PKR'
    
    return render_template('user/actual_timeline_input.html',
                         project=project,
                         monthly_data_for_actuals=monthly_data_for_actuals,
                         currency=currency,
                         max_actual_value=max_actual_value)


    
    
    
@app.route('/user/continue_project/<project_id>')
@login_required(role='user')
def continue_project(project_id):
    try:
        # Store project ID in session for continuation
        session['current_project_id'] = project_id
        return redirect(url_for('project_parameters', project_id=project_id))
    except Exception as e:
        flash('Error redirecting to project editor', 'danger')
        return redirect(url_for('user_dashboard'))
    
@app.route('/user/milestone_view/<project_id>')
@login_required(role='user')
def milestone_view(project_id):
    project = db.projects.find_one({'project_id': project_id, 'created_by_user_id': session['user_id']})
    if not project:
        flash('Project not found or you do not have access.', 'danger')
        return redirect(url_for('user_dashboard'))
    
    # Get currency and unit cost parameter from the database
    unit_costs_doc = db.unit_costs.find_one()
    currency = unit_costs_doc.get('currency', 'PKR') if unit_costs_doc else 'PKR'
    
    # Get the unit cost from project_parameters if available, otherwise use a default value
    unit_cost_per_unit = 50
    unit_type = "Units"
    if project.get('project_parameters') and len(project['project_parameters']) > 0:
        # Assuming the first parameter in the list contains the cost
        unit_cost_per_unit = project['project_parameters'][0].get('cost', 50)
        unit_type = project['project_parameters'][0].get('unit', "Units")
    
    # Initialize total variables
    total_forecasted_units = 0
    total_forecasted_cost_savings = 0
    total_actual_units = 0
    total_actual_cost_savings = 0

    # Process forecasted data - ensure all months are included with correct values
    forecasted_data = []
    for month_data in project.get('monthlyData', []):
        # Calculate total value - use monthly total if specified, otherwise sum daily values
        total_value = month_data.get('totalValue', 0)
        if not month_data.get('userEnteredMonthly', False):
            total_value = sum(day.get('value', 0) for day in month_data.get('days', []))
        
        # Calculate per cost savings for the month
        per_cost_savings = total_value * unit_cost_per_unit
        
        # Add to total sum for forecasted data
        total_forecasted_units += total_value
        total_forecasted_cost_savings += per_cost_savings
        
        # Process daily data to add per cost savings
        days_with_savings = []
        if not month_data.get('userEnteredMonthly', False):
            days_with_savings = [
                {
                    **day,
                    'per_cost_savings': day.get('value', 0) * unit_cost_per_unit
                } for day in month_data.get('days', [])
            ]
        
        forecasted_data.append({
            'month': month_data['month'],
            'year': month_data['year'],
            'totalValue': total_value,
            'per_cost_savings': per_cost_savings,
            'entryType': 'Monthly Total' if month_data.get('userEnteredMonthly', False) else 'By Day',
            'details': 'N/A' if month_data.get('userEnteredMonthly', False) else 'Daily Breakdown',
            'days': days_with_savings,
            'userEnteredMonthly': month_data.get('userEnteredMonthly', False)
        })
    
    # Process actual data similarly
    actual_data = []
    for month_data in project.get('actualMonthlyData', []):
        total_value = month_data.get('totalValue', 0)
        if not month_data.get('userEnteredMonthly', False):
            total_value = sum(day.get('value', 0) for day in month_data.get('days', []))
            
        # Calculate per cost savings for the month
        per_cost_savings = total_value * unit_cost_per_unit

        # Add to total sum for actual data
        total_actual_units += total_value
        total_actual_cost_savings += per_cost_savings

        # Process daily data to add per cost savings
        days_with_savings = []
        if not month_data.get('userEnteredMonthly', False):
            days_with_savings = [
                {
                    **day,
                    'per_cost_savings': day.get('value', 0) * unit_cost_per_unit
                } for day in month_data.get('days', [])
            ]

        actual_data.append({
            'month': month_data['month'],
            'year': month_data['year'],
            'totalValue': total_value,
            'per_cost_savings': per_cost_savings,
            'entryType': 'Monthly Total' if month_data.get('userEnteredMonthly', False) else 'By Day',
            'details': 'N/A' if month_data.get('userEnteredMonthly', False) else 'Daily Breakdown',
            'days': days_with_savings,
            'userEnteredMonthly': month_data.get('userEnteredMonthly', False)
        })
    
    return render_template('user/milestone_view.html', 
                           project=project, 
                           currency=currency,
                           forecasted_data=forecasted_data,
                           actual_data=actual_data,
                           unit_type=unit_type,
                           total_forecasted_units=total_forecasted_units,
                           total_forecasted_cost_savings=total_forecasted_cost_savings,
                           total_actual_units=total_actual_units,
                           total_actual_cost_savings=total_actual_cost_savings)

    
    

@app.route('/user/project_details/<project_id>')
@login_required(role='user')
def user_project_details(project_id):
    try:
        print(f"DEBUG: Starting project details for {project_id}")  # Debug point 1
        
        # Verify user session
        if 'user_id' not in session:
            flash('Please login to view projects', 'danger')
            return redirect(url_for('login'))
            
        print(f"DEBUG: User {session['user_id']} accessing project")  # Debug point 2

        # Fetch project
        project = db.projects.find_one({
            'project_id': project_id, 
            'created_by_user_id': session['user_id']
        })
        
        if not project:
            flash('Project not found or you do not have access.', 'danger')
            return redirect(url_for('user_dashboard'))
            
        print(f"DEBUG: Project found - {project.get('project_name')}")  # Debug point 3

        # Get currency info
        unit_costs_doc = db.unit_costs.find_one()
        currency = unit_costs_doc.get('currency', 'PKR') if unit_costs_doc else 'PKR'
        
        # Date handling with error protection
        date_fields = ['created_at', 'updated_at', 'start_date', 'end_date', 'reviewed_at']
        for field in date_fields:
            if field in project and isinstance(project[field], str):
                try:
                    if field in ['start_date', 'end_date']:
                        project[field] = datetime.strptime(project[field], '%Y-%m-%d')
                    else:
                        project[field] = datetime.fromisoformat(project[field])
                except ValueError as e:
                    print(f"WARNING: Failed to parse {field}: {str(e)}")
                    project[field] = None

        # ROI data handling
        if 'roi_data' in project and isinstance(project['roi_data'].get('calculated_at'), str):
            try:
                project['roi_data']['calculated_at'] = datetime.fromisoformat(project['roi_data']['calculated_at'])
            except (ValueError, AttributeError) as e:
                print(f"WARNING: Failed to parse ROI calculated_at: {str(e)}")

        print("DEBUG: All data processed successfully")  # Debug point 4
        
        return render_template('user/project_details.html',  # Double check this path matches your actual template
                            project=project,
                            currency=currency,
                            display_parameters=project.get('project_parameters', []),
                            machinery_cost=project.get('machinery_cost', 0),
                            dmaic_phases=["Define", "Measure", "Analyze", "Improve", "Control"])
                            
    except Exception as e:
        error_msg = f"Error loading project: {str(e)}"
        print(error_msg)
        print(traceback.format_exc())  # This will print the full traceback to your console
        flash('An error occurred while loading project details', 'danger')
        return redirect(url_for('user_dashboard'))
    
    
@app.route('/user/roi_table/<project_id>', methods=['GET', 'POST'])
@login_required(role='user')
def roi_table(project_id):
    try:
        project = db.projects.find_one({
            'project_id': project_id,
            'created_by_user_id': session['user_id']
        })
        if not project:
            flash('Project not found or you do not have access.', 'danger')
            return redirect(url_for('user_dashboard'))

        # Handle years_to_show instead of months_to_show
        try:
            years_to_show = int(request.form.get('years', 5))  # Default to 5 years
            years_to_show = max(1, years_to_show)
        except ValueError:
            years_to_show = 5

        # Calculate ROI - pass project data as a dict for calculate_roi
        # We'll calculate monthly first then convert to annual
        months_to_calculate = years_to_show * 12
        roi_data = calculate_roi({
            'project_type': project.get('project_type'),
            'project_parameters': project.get('project_parameters', []),
            'machinery_cost': project.get('machinery_cost')
        }, months_to_calculate)

        # Convert monthly data to annual
        annual_data = {
            'is_kaizen': roi_data.get('is_kaizen', False),
            'machinery_cost': roi_data.get('machinery_cost', 0),
            'annual_saving': roi_data.get('monthly_saving', 0) * 12,
            'roi_year': None,
            'years': []
        }

        # Group months into years
        for year_num in range(1, years_to_show + 1):
            start_month = (year_num - 1) * 12
            end_month = year_num * 12
            year_months = roi_data['months'][start_month:end_month]
            
            if not year_months:
                continue
                
            annual_saving = sum(month['saving'] for month in year_months)
            cumulative_saving = year_months[-1]['cumulative']
            is_roi = any(month['is_roi'] for month in year_months)
            
            if is_roi and annual_data['roi_year'] is None:
                annual_data['roi_year'] = year_num
                
            annual_data['years'].append({
                'year': year_num,
                'saving': annual_saving,
                'cumulative': cumulative_saving,
                'is_roi': is_roi
            })

        # Fetch currency
        unit_costs_doc = db.unit_costs.find_one({}, {'currency': 1})
        currency = unit_costs_doc.get('currency', 'PKR') if unit_costs_doc else 'PKR'

        return render_template(
            'user/roi_table.html',
            project=project,
            roi_data=annual_data,
            years_to_show=years_to_show,
            currency=currency
        )
    except Exception as e:
        flash(f'An error occurred: {str(e)}', 'danger')
        return redirect(url_for('user_dashboard'))
@app.route('/export-roi/<string:project_id>')
@login_required
def export_roi(project_id):
    project = db.projects.find_one({'project_id': project_id})
    if not project:
        flash('Project not found.', 'danger')
        return redirect(url_for('user_dashboard'))
    roi_data = project.get('roi_data')
    if not roi_data:
        # Recalculate if not found (e.g., for older projects)
        roi_data = calculate_roi({
            'project_type': project.get('project_type'),
            'project_parameters': project.get('project_parameters', []),
            'machinery_cost': project.get('machinery_cost')
        }, months_to_show=60)
    filename = secure_filename(f"ROI_Report_{project.get('project_id', 'N/A')}.xlsx")
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    try:
        export_roi_to_excel(project, roi_data, filepath)
        flash('ROI report generated successfully!', 'success')
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        flash(f'Error generating Excel report: {e}', 'danger')
        return redirect(url_for('user_project_details', project_id=project_id))

@app.route('/export_pdf/<project_id>')
@login_required()
def export_pdf(project_id):
    try:
        filepath = generate_report(project_id)
        flash('PDF report generated successfully!', 'success')
        return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))
    except Exception as e:
        flash(f'Error generating PDF report: {e}', 'danger')
        return redirect(url_for('user_project_details', project_id=project_id))

@app.route('/user/update_six_sigma_phase/<project_id>', methods=['POST'])
@login_required(role='user')
def update_six_sigma_phase(project_id):
    project = db.projects.find_one({'project_id': project_id, 'created_by_user_id': session['user_id']})
    if not project:
        flash('Project not found or you do not have access.', 'danger')
        return redirect(url_for('user_dashboard'))
    
    # Only allow phase updates for approved Six Sigma projects
    if project.get('project_type') != 'Six Sigma' or project.get('status') != 'Approved':
        flash('Six Sigma phases can only be updated for approved Six Sigma projects.', 'warning')
        return redirect(url_for('user_project_details', project_id=project_id))
    
    new_phase = request.form.get('six_sigma_phase')
    dmaic_phases = ["Define", "Measure", "Analyze", "Improve", "Control"]
    if new_phase and new_phase in dmaic_phases:
        try:
            db.projects.update_one(
                {'_id': project['_id']},
                {'$set': {'six_sigma_phase': new_phase, 'updated_at': datetime.now()}}
            )
            flash(f'Six Sigma phase updated to "{new_phase}" successfully!', 'success')
        except Exception as e:
            flash(f'Error updating Six Sigma phase: {e}', 'danger')
    else:
        flash('Invalid Six Sigma phase selected.', 'danger')
    return redirect(url_for('user_project_details', project_id=project_id))


# Currency API Routes
@app.route('/api/currencies', methods=['GET', 'POST'])
@login_required(role='admin')
def api_currencies():
    if request.method == 'GET':
        # Get all currencies
        currencies = list(db.currency_rates.find().sort([('currency_name', 1), ('year', -1)]))
        return jsonify([{
            'currency_name': c['currency_name'],
            'year': c['year'],
            'monthly_rates': c['monthly_rates'],
            'created_at': c['created_at'].isoformat() if 'created_at' in c else None
        } for c in currencies])
    
    elif request.method == 'POST':
        # Add/update currency rates
        data = request.get_json()
        currency_name = data.get('currency_name')
        year = data.get('year')
        rates = data.get('rates')
        
        if not all([currency_name, year, rates]):
            return jsonify({'error': 'Missing required fields'}), 400
        
        # Calculate YTD average (only for entered months)
        entered_rates = [v for v in rates.values() if v is not None]
        ytd_avg = sum(entered_rates) / len(entered_rates) if entered_rates else 0
        
        # Create/update record
        db.currency_rates.update_one(
            {'currency_name': currency_name, 'year': year},
            {'$set': {
                'monthly_rates': rates,
                'ytd_avg': ytd_avg,
                'updated_at': datetime.now()
            }},
            upsert=True
        )
        
        return jsonify({'success': True})

@app.route('/api/currency/<currency_name>/<int:year>', methods=['GET', 'DELETE'])
@login_required(role='admin')
def api_currency(currency_name, year):
    if request.method == 'GET':
        currency = db.currency_rates.find_one({'currency_name': currency_name, 'year': year})
        if not currency:
            return jsonify({'error': 'Currency not found'}), 404
        return jsonify({
            'currency_name': currency['currency_name'],
            'year': currency['year'],
            'monthly_rates': currency['monthly_rates'],
            'created_at': currency.get('created_at', datetime.now()).isoformat()
        })
    
    elif request.method == 'DELETE':
        result = db.currency_rates.delete_one({'currency_name': currency_name, 'year': year})
        if result.deleted_count == 0:
            return jsonify({'error': 'Currency not found'}), 404
        return jsonify({'success': True})

@app.route('/api/convert', methods=['GET'])
@login_required()
def api_convert():
    currency = request.args.get('currency')
    month = request.args.get('month')
    year = request.args.get('year', datetime.now().year)
    
    if not currency or not month:
        return jsonify({'error': 'Currency and month are required'}), 400
    
    # Get currency data
    currency_data = db.currency_rates.find_one({
        'currency_name': currency,
        'year': int(year)
    })
    
    if not currency_data:
        return jsonify({'error': 'Currency data not found'}), 404
    
    # Calculate YTD up to selected month
    months = ['January', 'February', 'March', 'April', 'May', 'June', 
             'July', 'August', 'September', 'October', 'November', 'December']
    month_index = months.index(month)
    ytd_months = months[:month_index + 1]
    
    rates = [currency_data['monthly_rates'][m] for m in ytd_months if currency_data['monthly_rates'][m] is not None]
    if not rates:
        return jsonify({'error': 'No rates available for selected period'}), 400
    
    ytd_value = sum(rates) / len(rates)
    current_rate = currency_data['monthly_rates'][month]
    
    return jsonify({
        'rate': current_rate,
        'ytd': ytd_value,
        'currency': currency,
        'month': month,
        'year': year
    })

# Admin Routes
@app.route('/admin/dashboard')
@login_required(role='admin')
def admin_dashboard():
    total_users = db.users.count_documents({})
    total_projects = db.projects.count_documents({})
    pending_projects = db.projects.count_documents({'status': 'Pending'})
    approved_projects = db.projects.count_documents({'status': 'Approved'})
    recent_projects = list(db.projects.find().sort('created_at', -1).limit(5))
    return render_template('admin/admin_dashboard.html',
                           total_users=total_users,
                           total_projects=total_projects,
                           pending_projects=pending_projects,
                           approved_projects=approved_projects,
                           recent_projects=recent_projects)

@app.route('/admin/projects')
@login_required(role='admin')
def admin_projects():
    # Initialize query
    query = {}
    
    # Get filter parameters from request
    project_type_filter = request.args.get('project_type_filter', 'All')
    year_filter = request.args.get('year_filter', 'All')
    month_filter = request.args.get('month_filter', 'All')
    
    # Apply project type filter
    if project_type_filter and project_type_filter != 'All':
        query['project_type'] = project_type_filter
    
    # Get all projects (we'll filter dates in Python for flexibility)
    projects = list(db.projects.find(query).sort('created_at', -1))
    
    # Process date filters
    filtered_projects = []
    all_years = set()
    for project in projects:
        # Extract year from created_at
        if 'created_at' in project and isinstance(project['created_at'], datetime):
            project_year = project['created_at'].year
            project_month = project['created_at'].month
            all_years.add(project_year)
            
            # Apply year filter if specified
            if year_filter != 'All' and str(project_year) != year_filter:
                continue
                
            # Apply month filter if specified
            if month_filter != 'All' and str(project_month) != month_filter:
                continue
                
            filtered_projects.append(project)
        else:
            # Include projects without dates if no date filter is applied
            if year_filter == 'All' and month_filter == 'All':
                filtered_projects.append(project)
    
    # Prepare months data (using a list of dicts for consistency)
    all_months = [
        {'value': '1', 'name': 'January'},
        {'value': '2', 'name': 'February'},
        {'value': '3', 'name': 'March'},
        {'value': '4', 'name': 'April'},
        {'value': '5', 'name': 'May'},
        {'value': '6', 'name': 'June'},
        {'value': '7', 'name': 'July'},
        {'value': '8', 'name': 'August'},
        {'value': '9', 'name': 'September'},
        {'value': '10', 'name': 'October'},
        {'value': '11', 'name': 'November'},
        {'value': '12', 'name': 'December'}
    ]
    
    # Get project types
    project_types = list(db.project_categories.find().sort('name', 1))
    
    # Get currency
    unit_costs_doc = db.unit_costs.find_one()
    currency = unit_costs_doc.get('currency', 'PKR') if unit_costs_doc else 'PKR'
    
    return render_template('admin/projects.html',
                         projects=filtered_projects,
                         project_types=project_types,
                         all_years=sorted(all_years, reverse=True),
                         all_months=all_months,
                         selected_project_type=project_type_filter,
                         selected_year=year_filter,
                         selected_month=month_filter,
                         currency=currency)



    
@app.route('/export-roi/<string:project_id>', endpoint='export_roi')
@login_required
def export_roi(project_id):
    # Changed to use MongoDB db.projects collection
    project = db.projects.find_one({'project_id': project_id})
    if not project:
        flash('Project not found.', 'danger')
        return redirect(url_for('user_dashboard'))
    roi_data = project.get('roi_data')
    if not roi_data:
        # Recalculate if not found (e.g., for older projects)
        roi_data = calculate_roi({
            'project_type': project.get('project_type'),
            'project_parameters': project.get('project_parameters', []),
            'machinery_cost': project.get('machinery_cost')
        }, months_to_show=60)
    filename = secure_filename(f"ROI_Report_{project.get('project_id', 'N/A')}.xlsx")
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    try:
        export_roi_to_excel(project, roi_data, filepath)
        flash('ROI report generated successfully!', 'success')
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        flash(f'Error generating PDF report: {e}', 'danger')
        return redirect(url_for('user_project_details', project_id=project_id))

@app.route('/admin/project_detail/<project_id>')
@login_required(role='admin')
def admin_project_detail(project_id):
    project = db.projects.find_one({'project_id': project_id})
    if not project:
        flash('Project not found.', 'danger')
        return redirect(url_for('admin_projects'))
    
    # Get currency
    unit_costs_doc = db.unit_costs.find_one()
    currency = unit_costs_doc.get('currency', 'PKR') if unit_costs_doc else 'PKR'
    
    # Convert string dates to datetime objects
    date_fields = ['created_at', 'updated_at', 'start_date', 'end_date', 'reviewed_at']
    for field in date_fields:
        if field in project and isinstance(project[field], str):
            try:
                if field in ['start_date', 'end_date']:
                    project[field] = datetime.strptime(project[field], '%Y-%m-%d')
                else:
                    project[field] = datetime.fromisoformat(project[field])
            except ValueError:
                project[field] = None
    
    # Handle ROI data date conversion
    if project.get('roi_data') and isinstance(project['roi_data'].get('calculated_at'), str):
        try:
            project['roi_data']['calculated_at'] = datetime.fromisoformat(project['roi_data']['calculated_at'])
        except ValueError:
            project['roi_data']['calculated_at'] = None
    
    return render_template(
        'admin/project_detail.html',
        project=project,
        currency=currency,
        display_parameters=project.get('project_parameters', []),
        machinery_cost=project.get('machinery_cost', 0)
    )

@app.route('/admin/manage_accounts/<role_type>')
@login_required(role='admin')
def admin_manage_accounts(role_type):
    collection = get_user_collection(role_type)
    if collection is None:
        flash("Invalid role type.", "danger")
        return redirect(url_for('admin_dashboard'))
    users = list(collection.find({}).sort('created_at', -1))
    return render_template('admin/manage_accounts.html', users=users, role_type=role_type)

@app.route('/admin/add_account', methods=['GET', 'POST'])
@login_required(role='admin')
def admin_add_account():
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        role = request.form['role']
        status = int(request.form.get('status', 0))
        collection = get_user_collection(role)
        if collection is None:
            flash("Invalid role selected.", "danger")
            return redirect(url_for('admin_add_account'))
        if db.users.find_one({'email': email}) or \
           db.managers.find_one({'email': email}) or \
           db.admins.find_one({'email': email}):
            flash('Email already registered for another account.', 'danger')
            return render_template('admin/add_edit_account.html', edit_mode=False, account={
                'username': username, 'email': email, 'role': role, 'status': status
            })
        try:
            hashed_password = generate_password_hash(password)
            new_account = {
                'username': username,
                'email': email,
                'password': hashed_password,
                'role': role,
                'status': status,
                'created_at': datetime.now()
            }
            collection.insert_one(new_account)
            flash(f'{role.capitalize()} account "{username}" created successfully!', 'success')
            return redirect(url_for('admin_manage_accounts', role_type=role))
        except Exception as e:
            flash(f'Error creating account: {e}', 'danger')
            return render_template('admin/add_edit_account.html', edit_mode=False, account={
                'username': username, 'email': email, 'role': role, 'status': status
            })
    return render_template('admin/add_edit_account.html', edit_mode=False, account={})

@app.route('/admin/edit_account/<role_type>/<user_id>', methods=['GET', 'POST'])
@login_required(role='admin')
def admin_edit_account(role_type, user_id):
    collection = get_user_collection(role_type)
    if collection is None:
        flash("Invalid role type.", "danger")
        return redirect(url_for('admin_dashboard'))
    account = collection.find_one({'_id': ObjectId(user_id)})
    if not account:
        flash("Account not found.", "danger")
        return redirect(url_for('admin_manage_accounts', role_type=role_type))
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form.get('password')
        new_role = request.form['role']
        status = int(request.form.get('status', 0))
        if email != account['email']:
            if db.users.find_one({'email': email}) or \
               db.managers.find_one({'email': email}) or \
               db.admins.find_one({'email': email}):
                flash('Email already registered for another account.', 'danger')
                return render_template('admin/add_edit_account.html', edit_mode=True, account=account)
        update_data = {
            'username': username,
            'email': email,
            'role': new_role,
            'status': status,
            'updated_at': datetime.now()
        }
        if password:
            update_data['password'] = generate_password_hash(password)
        try:
            if new_role != role_type:
                new_collection = get_user_collection(new_role)
                if new_collection is None:
                    flash("Invalid new role selected.", "danger")
                    return render_template('admin/add_edit_account.html', edit_mode=True, account=account)
                new_collection.insert_one(update_data)
                collection.delete_one({'_id': ObjectId(user_id)})
                flash(f'Account "{username}" moved to {new_role.capitalize()}s successfully!', 'success')
                return redirect(url_for('admin_manage_accounts', role_type=new_role))
            else:
                collection.update_one({'_id': ObjectId(user_id)}, {'$set': update_data})
                flash(f'{role_type.capitalize()} account "{username}" updated successfully!', 'success')
                return redirect(url_for('admin_manage_accounts', role_type=role_type))
        except Exception as e:
            flash(f'Error updating account: {e}', 'danger')
            return render_template('admin/add_edit_account.html', edit_mode=True, account=account)
    return render_template('admin/add_edit_account.html', edit_mode=True, account=account)




@app.route('/admin/project_categories')
@login_required(role='admin')
def admin_project_categories():
    categories = list(db.project_categories.find().sort('name', 1))
    return render_template('admin/project_category.html', categories=categories)


@app.route('/admin/toggle-category-status/<category_id>', methods=['POST'])
@login_required(role='admin')
def admin_toggle_category_status(category_id):
    try:
        # Get current status
        category = db.project_categories.find_one({'_id': ObjectId(category_id)})
        if not category:
            flash('Category not found', 'danger')
            return redirect(url_for('admin_project_categories'))
        
        # Toggle the status
        new_status = not category.get('is_active', True)
        
        # Update in database
        result = db.project_categories.update_one(
            {'_id': ObjectId(category_id)},
            {'$set': {'is_active': new_status}}
        )
        
        if result.modified_count > 0:
            flash(f'Category {"activated" if new_status else "deactivated"} successfully', 'success')
        else:
            flash('No changes made to category status', 'info')
            
    except Exception as e:
        flash(f'Error updating category status: {str(e)}', 'danger')
    
    return redirect(url_for('admin_project_categories'))


@app.route('/admin/add_project_category', methods=['POST'])
@login_required(role='admin')
def admin_add_project_category():
    category_name = request.form['category_name'].strip()
    if not category_name:
        flash('Category name cannot be empty.', 'danger')
        return redirect(url_for('admin_project_categories'))
    if db.project_categories.find_one({'name': category_name}):
        flash(f'Category "{category_name}" already exists.', 'warning')
        return redirect(url_for('admin_project_categories'))
    try:
        db.project_categories.insert_one({'name': category_name, 'created_at': datetime.now()})
        flash(f'Category "{category_name}" added successfully!', 'success')
    except Exception as e:
        flash(f'Error adding category: {e}', 'danger')
    return redirect(url_for('admin_project_categories'))

@app.route('/admin/edit_project_category/<category_id>', methods=['POST'])
@login_required(role='admin')
def admin_edit_project_category(category_id):
    category_name = request.form['category_name'].strip()
    if not category_name:
        flash('Category name cannot be empty.', 'danger')
        return redirect(url_for('admin_project_categories'))
    if db.project_categories.find_one({'name': category_name, '_id': {'$ne': ObjectId(category_id)}}):
        flash(f'Category "{category_name}" already exists.', 'warning')
        return redirect(url_for('admin_project_categories'))
    try:
        result = db.project_categories.update_one(
            {'_id': ObjectId(category_id)},
            {'$set': {'name': category_name, 'updated_at': datetime.now()}}
        )
        if result.modified_count > 0:
            flash(f'Category updated to "{category_name}" successfully!', 'success')
        else:
            flash('Category not found or no changes made.', 'info')
    except Exception as e:
        flash(f'Error updating category: {e}', 'danger')
    return redirect(url_for('admin_project_categories'))

@app.route('/admin/upload', methods=['GET', 'POST'])
@login_required(role='admin')
def admin_upload():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('No selected file', 'danger')
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            try:
                filename = secure_filename(file.filename)
                temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"temp_{filename}")
                file.save(temp_filepath)
                
                # Quick validation before processing
                if filename.endswith('.csv'):
                    df = pd.read_csv(temp_filepath, nrows=1)
                else:
                    df = pd.read_excel(temp_filepath, nrows=1)
                
                required_columns = ['Plant_Number']
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    os.remove(temp_filepath)
                    flash(f'Missing required columns: {", ".join(missing_columns)}', 'danger')
                    return redirect(request.url)
                
                # Process the full file
                success, message, processed_files = process_factory_data(temp_filepath)
                
                if success:
                    flash(f'Success: {message}', 'success')
                    if db is not None:
                        db.reports.insert_one({
                            "type": "data_upload",
                            "filename": filename,
                            "processed_files": processed_files,
                            "status": "success",
                            "message": message,
                            "timestamp": datetime.now(),
                            "uploaded_by": session.get('user_id'),
                            "uploaded_by_username": session.get('username')
                        })
                else:
                    flash(f'Error: {message}', 'danger')
                    if db is not None:
                        db.reports.insert_one({
                            "type": "data_upload",
                            "filename": filename,
                            "status": "failed",
                            "error": message,
                            "timestamp": datetime.now(),
                            "uploaded_by": session.get('user_id'),
                            "uploaded_by_username": session.get('username')
                        })
                
                if os.path.exists(temp_filepath):
                    os.remove(temp_filepath)
                return redirect(url_for('admin_upload'))
                
            except Exception as e:
                if 'temp_filepath' in locals() and os.path.exists(temp_filepath):
                    os.remove(temp_filepath)
                flash(f'Upload error: {str(e)}', 'danger')
                return redirect(request.url)
        else:
            flash('Invalid file type. Please upload an Excel (.xlsx) or CSV (.csv) file.', 'danger')
            return redirect(request.url)
    
    upload_history = []
    if db is not None:
        upload_history = list(db.reports.find(
            {"type": "data_upload"},
            {"filename": 1, "status": 1, "timestamp": 1, "processed_files": 1, "uploaded_by_username": 1}
        ).sort("timestamp", -1).limit(10))
    
    return render_template('admin/upload.html', upload_history=upload_history)

@app.route('/download-template')
@login_required(role='admin')
def download_template():
    try:
        # Simple direct path to your file
        return send_from_directory(
            directory=os.path.join(app.static_folder, 'exports'),
            path='factory_data_template.csv',
            as_attachment=True,
            mimetype='text/csv'
        )
    except Exception as e:
        flash("Failed to download template. Please try again.", "danger")
        return redirect(url_for('admin_upload'))
@app.route('/admin/create_project', methods=['GET', 'POST'])
@login_required(role='admin')
def admin_create_project():
    all_factories = list(db.factories.find({}))
    all_products = {}
    all_models = {}
    for factory_doc in all_factories:
        factory_code = factory_doc.get('factory_code')
        for p_data in factory_doc.get('products', []):
            p_data_lower_keys = {k.lower(): v for k, v in p_data.items()}
            product_name = None
            for key_candidate in ['product_name', 'product', 'name', 'item', 'description']:
                if key_candidate in p_data_lower_keys and p_data_lower_keys[key_candidate] is not None:
                    product_name_candidate = str(p_data_lower_keys[key_candidate]).strip()
                    if product_name_candidate:
                        product_name = product_name_candidate
                        break
            if product_name:
                if product_name not in all_products:
                    all_products[product_name] = {'product_id': product_name, 'name': product_name, 'factory_code': factory_code}
            model_name = None
            for key_candidate in ['model', 'variant', 'sku', 'id', 'name']:
                if key_candidate in p_data_lower_keys and p_data_lower_keys[key_candidate] is not None:
                    model_name_candidate = str(p_data_lower_keys[key_candidate]).strip()
                    if model_name_candidate:
                        model_name = model_name_candidate
                        break
            if model_name:
                if model_name not in all_models:
                    all_models[model_name] = {'model_id': model_name, 'name': model_name, 'product_id': product_name}
    products = list(all_products.values())
    models = list(all_models.values())

    unit_costs_doc = db.unit_costs.find_one()
    currency = unit_costs_doc.get('currency', 'PKR') if unit_costs_doc else 'PKR'
    unit_costs_for_display = unit_costs_doc.get('DPL1', []) if unit_costs_doc else []
    unit_cost_lookup = {uc['name']: {'unit': uc['unit'], 'cost': uc['cost'], 'operator': uc['operator']} for uc in unit_costs_for_display}

    project_data = {}
    existing_param_names = [p['name'] for p in project_data.get('project_parameters', [])]
    project_categories = list(db.project_categories.find().sort('name', 1))

    if request.method == 'POST':
        project_name = request.form['project_name']
        project_description = request.form['project_description']
        project_type = request.form['project_type']
        factory_code = request.form['factory_code']
        start_date_str = request.form['start_date']
        end_date_str = request.form['end_date']
        selected_products = request.form.getlist('selected_products')
        selected_models = request.form.getlist('selected_models')
        param_names = request.form.getlist('param_name[]')
        param_units = request.form.getlist('param_unit[]')
        param_costs = request.form.getlist('param_cost[]')
        param_before_values = request.form.getlist('param_before_value[]')
        param_after_values = request.form.getlist('param_after_values[]')
        param_operators = request.form.getlist('param_operator[]')
        
        project_parameters = []
        for i in range(len(param_names)):
            try:
                name = param_names[i].strip()
                before_value = float(param_before_values[i] or 0.0)
                after_value = float(param_after_values[i] or 0.0)
                operator = param_operators[i].strip()
                is_custom_param = request.form.get(f'is_custom_param_{i}', 'false') == 'true'
                unit = param_units[i].strip()
                cost = float(param_costs[i] or 0.0)
                if not is_custom_param and name in unit_cost_lookup:
                    predefined_info = unit_cost_lookup[name]
                    unit = predefined_info['unit']
                    cost = predefined_info['cost']
                    operator = predefined_info['operator']
                project_parameters.append({
                    'name': name,
                    'unit': unit,
                    'cost': cost,
                    'before_value': before_value,
                    'after_value': after_value,
                    'operator': operator,
                    'is_custom': is_custom_param
                })
            except ValueError as e:
                flash(f"Invalid number format for parameter '{param_names[i]}'. Please enter valid numbers. Error: {e}", 'danger')
                project_data = {
                    'project_name': project_name,
                    'project_description': project_description,
                    'project_type': project_type,
                    'factory_code': factory_code,
                    'start_date': start_date_str,
                    'end_date': end_date_str,
                    'selected_products': selected_products,
                    'selected_models': selected_models,
                    'project_parameters': project_parameters,
                    'machinery_cost': request.form.get('machinery_cost', 0)
                }
                return render_template('admin/create_project.html', products=products, models=models, unit_costs=unit_costs_for_display, currency=currency, project_data=project_data, all_factories=all_factories, project_categories=project_categories, existing_param_names=[p['name'] for p in project_parameters])
        
        machinery_cost = float(request.form.get('machinery_cost', 0.0) or 0.0)
        
        temp_project_data = {
            'project_type': project_type,
            'project_parameters': project_parameters,
            'machinery_cost': machinery_cost
        }
        roi_result = calculate_roi(temp_project_data, months_to_show=60)
        
        start_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date_str, '%Y-%m-%d')
        monthly_data = []
        current_month_dt = start_dt.replace(day=1)
        while current_month_dt <= end_dt:
            month_name = current_month_dt.strftime('%B %Y')
            year = current_month_dt.year
            month_number = current_month_dt.month
            days_in_month = calendar.monthrange(year, month_number)[1]
            days_list = []
            for day_num in range(1, days_in_month + 1):
                days_list.append({'day': float(day_num), 'value': 0.0, 'notes': ''})
            monthly_data.append({
                'month': month_name,
                'year': year,
                'monthNumber': month_number,
                'days': days_list,
                'totalValue': roi_result.get('monthly_saving', 0.0),
                'userEnteredMonthly': True
            })
            if current_month_dt.month == 12:
                current_month_dt = current_month_dt.replace(year=current_month_dt.year + 1, month=1)
            else:
                current_month_dt = current_month_dt.replace(month=current_month_dt.month + 1)
        
        # Create new Project document for MongoDB
        project_id = f'PROJ-{db.projects.count_documents({}) + 1:04d}'
        new_project = {
            'project_id': project_id,
            'project_name': project_name,
            'project_description': project_description,
            'project_type': project_type,
            'factory_code': factory_code,
            'start_date': start_date_str, # Store as string
            'end_date': end_date_str,     # Store as string
            'selected_products': selected_products,
            'selected_models': selected_models,
            'project_parameters': project_parameters,
            'machinery_cost': machinery_cost,
            'monthlyData': monthly_data,
            'actualMonthlyData': [{
                'month': month['month'],
                'year': month['year'],
                'monthNumber': month['monthNumber'],
                'days': [{'day': day['day'], 'value': 0.0, 'notes': ''} for day in month.get('days', [])],
                'totalValue': 0.0,
                'userEnteredMonthly': month['userEnteredMonthly']
            } for month in monthly_data],
            'roi_data': roi_result,
            'estimated_savings': roi_result.get('monthly_saving', 0.0),
            'roi': roi_result.get('roi_month', 0.0), # Using roi_month as a proxy for ROI percentage if needed
            'payback_period': roi_result.get('roi_month', 0.0), # Using roi_month as payback period in months
            'created_by_user_id': session['user_id'],
            'created_by_username': session['username'],
            'created_at': datetime.now(),
            'status': 'Approved',
            'manager_comments': 'Approved by Admin during creation.',
            'reviewed_at': datetime.now(),
            'reviewed_by_user_id': session['user_id'],
            'reviewed_by_username': session['username']
        }
        if project_type == 'Six Sigma':
            new_project['six_sigma_phase'] = 'Define'
        db.projects.insert_one(new_project)
        flash('Project created and approved successfully!', 'success')
        return redirect(url_for('admin_project_detail', project_id=project_id))

    return render_template('admin/create_project.html', products=products, models=models, unit_costs_for_display=unit_costs_for_display, currency=currency, project_data=project_data, all_factories=all_factories, project_categories=project_categories, existing_param_names=existing_param_names, unit_costs_doc=unit_costs_doc)

@app.route('/manager/export_all_projects')
@login_required(role='manager')
def manager_export_all_projects():
    projects = list(db.projects.find().sort('created_at', -1))
    if not projects:
        flash('No projects to export.', 'info')
        return redirect(url_for('manager_projects'))
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "All Projects"
    headers = [
        "Project ID", "Project Name", "Type", "Factory", "Status", "Created By",
        "Created At", "Start Date", "End Date", "Monthly Saving", "Machinery Cost",
        "ROI Month", "Six Sigma Phase", "Manager Comments", "Reviewed By", "Reviewed At"
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='800000', end_color='800000', fill_type='solid')
    header_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
    data_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='000000')
    )
    for project in projects:
        row_data = [
            project.get('project_id', 'N/A'),
            project.get('project_name', 'N/A'),
            project.get('project_type', 'N/A'),
            project.get('factory_code', 'N/A'),
            project.get('status', 'N/A'),
            project.get('created_by_username', 'N/A'),
            project.get('created_at').strftime('%Y-%m-%d %H:%M') if isinstance(project.get('created_at'), datetime) else 'N/A',
            project.get('start_date', 'N/A'),
            project.get('end_date', 'N/A'),
            project.get('roi_data', {}).get('monthly_saving', 0),
            project.get('roi_data', {}).get('machinery_cost', 0),
            project.get('roi_data', {}).get('roi_month', 'N/A'),
            project.get('six_sigma_phase', 'N/A'),
            project.get('manager_comments', 'N/A'),
            project.get('reviewed_by_username', 'N/A'),
            project.get('reviewed_at').strftime('%Y-%m-%d %H:%M') if isinstance(project.get('reviewed_at'), datetime) else 'N/A'
        ]
        current_row = sheet.max_row + 1
        for col_num, value in enumerate(row_data, 1):
            cell = sheet.cell(row=current_row, column=col_num, value=value)
            cell.border = data_border
            if col_num in [10, 11]:
                cell.font = Font(bold=True)
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width
    workbook.save(output)
    output.seek(0)
    filename = secure_filename(f"All_Projects_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/unit_costs', methods=['GET', 'POST'])
@login_required(role='admin')
def admin_unit_costs():
    unit_costs_doc = db.unit_costs.find_one()
    currency = unit_costs_doc.get('currency', 'PKR') if unit_costs_doc else 'PKR'
    unit_costs_data = {
        "DPL1": unit_costs_doc.get("DPL1", []) if unit_costs_doc else [],
        "DPL2": unit_costs_doc.get("DPL2", []) if unit_costs_doc else [],
        "URIL": unit_costs_doc.get("URIL", []) if unit_costs_doc else []
    }
    
    if request.method == 'POST':
        factory_code = request.form.get('factory_code')
        param_names = request.form.getlist('param_name[]')
        param_units = request.form.getlist('param_unit[]')
        param_costs = request.form.getlist('param_cost[]')
        param_operators = request.form.getlist('param_operator[]')
        
        updated_params = []
        for i in range(len(param_names)):
            try:
                cost = float(param_costs[i] or 0.0)
                operator = param_operators[i]
                updated_params.append({
                    'name': param_names[i],
                    'unit': param_units[i],
                    'cost': cost,
                    'operator': operator
                })
            except ValueError:
                flash(f"Invalid number format for cost of '{param_names[i]}'. Please enter a valid number.", 'danger')
                return redirect(url_for('admin_unit_costs'))
        
        if unit_costs_doc:
            db.unit_costs.update_one(
                {'_id': unit_costs_doc['_id']},
                {'$set': {factory_code: updated_params}}
            )
        else:
            new_doc = {
                "DPL1": [], "DPL2": [], "URIL": [],
                "currency": "PKR"
            }
            new_doc[factory_code] = updated_params
            db.unit_costs.insert_one(new_doc)
        
        flash(f'Unit costs for {factory_code} updated successfully!', 'success')
        return redirect(url_for('admin_unit_costs'))
    
    return render_template('admin/unit_costs.html', unit_costs=unit_costs_data, currency=currency)


@app.route('/admin/toggle_account_status/<role_type>/<user_id>', methods=['POST'])
@login_required(role='admin')
def admin_toggle_account_status(role_type, user_id):
    collection = get_user_collection(role_type)
    if collection is None:
        flash("Invalid role type.", "danger")
        return redirect(url_for('admin_dashboard'))
    account = collection.find_one({'_id': ObjectId(user_id)})
    if not account:
        flash("Account not found.", "danger")
        return redirect(url_for('admin_manage_accounts', role_type=role_type))
    new_status = 1 if account.get('status') == 0 else 0
    try:
        collection.update_one({'_id': ObjectId(user_id)}, {'$set': {'status': new_status, 'updated_at': datetime.now()}})
        flash(f'Account "{account.get("username", account["email"])}" status toggled to {"Enabled" if new_status == 1 else "Disabled"}.', 'success')
    except Exception as e:
        flash(f'Error toggling account status: {e}', 'danger')
    return redirect(url_for('admin_manage_accounts', role_type=role_type))



    
@app.route('/projects/<project_id>/review', methods=['POST'])
def review_project(project_id):
    if request.method == 'POST':
        status = request.form.get('status')
        comments = request.form.get('comments')
        
        if not status:
            flash('Decision is required', 'danger')
            return redirect(url_for('project_review', project_id=project_id))
        
        # Update project in MongoDB
        try:
            db.projects.update_one(
                {'_id': ObjectId(project_id)},
                {'$set': {
                    'status': status,
                    'manager_comments': comments,
                    'reviewed_at': datetime.utcnow(),
                    'reviewed_by': current_user.id,
                    'reviewed_by_username': current_user.username
                }}
            )
            flash('Review submitted successfully', 'success')
        except Exception as e:
            flash('Failed to update project', 'danger')
        
        return redirect(url_for('manager_projects'))


@app.route('/export/project_report/<project_id>')
@login_required(role='manager')
def export_project_report(project_id):
    try:
        project = db.projects.find_one({'project_id': project_id})
        if not project:
            flash('Project not found.', 'danger')
            return redirect(url_for('manager_projects'))
            
        # Generate PDF report (you can use your existing generate_report function)
        filepath = generate_report(project_id)
        return send_file(filepath, as_attachment=True, download_name=f"Project_Report_{project_id}.pdf")
        
    except Exception as e:
        flash(f'Error generating report: {str(e)}', 'danger')
        return redirect(url_for('manager_project_detail', project_id=project_id))
    
@app.route('/manager/project_detail/<project_id>', methods=['GET', 'POST'])
@login_required(role='manager')
def manager_project_detail(project_id):
    try:
        project = db.projects.find_one({'project_id': project_id})
        if not project:
            flash('Project not found.', 'danger')
            return redirect(url_for('manager_projects'))

        if request.method == 'POST':
            status = request.form.get('status')
            comments = request.form.get('comments', '').strip()
            
            if not status or status not in ['Approved', 'Rejected']:
                flash('Please select a valid decision (Approve or Reject)', 'danger')
                return redirect(url_for('manager_project_detail', project_id=project_id))
            
            if not comments:
                flash('Comments are required for review', 'warning')
                return redirect(url_for('manager_project_detail', project_id=project_id))

            # Create the comment object first
            new_comment = {
                'content': comments,
                'author': session['username'],
                'timestamp': datetime.now(),
                'type': 'review'
            }

            # Perform the update with proper MongoDB update operators
            result = db.projects.update_one(
                {'project_id': project_id},
                {
                    '$set': {
                        'status': status,
                        'manager_comments': comments,
                        'reviewed_at': datetime.now(),
                        'reviewed_by_user_id': session['user_id'],
                        'reviewed_by_username': session['username']
                    },
                    '$push': {
                        'comments': new_comment
                    }
                }
            )
            
            if result.modified_count == 1:
                flash(f'Project {status.lower()} successfully!', 'success')
            else:
                flash('Failed to update project status', 'danger')
            
            return redirect(url_for('manager_project_detail', project_id=project_id))

        # GET request handling
        unit_costs_doc = db.unit_costs.find_one() or {}
        currency = unit_costs_doc.get('currency', 'PKR')

        # Get all available currencies for the converter
        current_year = datetime.now().year
        currencies = list(db.currency_rates.find(
            {'year': current_year}, 
            {'currency_name': 1, 'monthly_rates': 1}
        ).sort('currency_name', 1))

        # Calculate ROI based on project parameters
        total_parameter_savings = 0
        machinery_cost = project.get('machinery_cost', 0)  # Use actual value, no default
            
        for param in project.get('project_parameters', []):
            if param['operator'] == '>':
                savings = (param['before_value'] - param['after_value']) * param['cost']
            else:  # operator == '<'
                savings = (param['after_value'] - param['before_value']) * param['cost']
            total_parameter_savings += savings

        # Calculate monthly savings (total savings divided by 12)
        monthly_savings = total_parameter_savings / 12 if total_parameter_savings > 0 else 0
            
        # Calculate ROI month (payback period) - months to recover investment
        roi_month = None
        if machinery_cost > 0 and monthly_savings > 0:
            roi_month = ceil(machinery_cost / monthly_savings)

        # Calculate ROI percentage from parameters
        roi_percentage = (total_parameter_savings / machinery_cost) * 100 if machinery_cost > 0 else 0
            
        # Calculate actual savings from monthly data if available
        actual_savings = 0
        if 'actualMonthlyData' in project and project['actualMonthlyData']:
            actual_savings = sum(month.get('totalValue', 0) for month in project['actualMonthlyData'])

        # Add ROI data to project object
        project['roi'] = roi_percentage
        project['roi_month'] = roi_month
        project['calculated_savings'] = total_parameter_savings
        project['actual_savings'] = actual_savings

        # Create ROI data structure for the template
        roi_data = {
            'monthly_saving': monthly_savings,
            'roi_month': roi_month,
            'machinery_cost': machinery_cost,
            'total_savings': total_parameter_savings
        }
        project['roi_data'] = roi_data

        # Date handling
        today = datetime.now().date()
        start_date = None
        end_date = None
        days_remaining = None
        try:
            if 'start_date' in project:
                start_date = (datetime.strptime(project['start_date'], '%Y-%m-%d') 
                             if isinstance(project['start_date'], str) 
                             else project['start_date']).date()
                
            if 'end_date' in project:
                end_date = (datetime.strptime(project['end_date'], '%Y-%m-%d') 
                          if isinstance(project['end_date'], str) 
                          else project['end_date']).date()
                if end_date and today:
                    days_remaining = max(0, (end_date - today).days)
        except (ValueError, TypeError) as e:
            flash('Error processing project dates', 'warning')
            logger.error(f"Date processing error for project {project_id}: {str(e)}")
            if 'comments' not in project:
                project['comments'] = []
        
        return render_template(
            'manager/project_detail.html',
            project=project,
            currency=currency,
            currencies=currencies,
            days_remaining=days_remaining,
            start_date=start_date,
            end_date=end_date,
            today=today,
            display_parameters=project.get('project_parameters', []),
            machinery_cost=machinery_cost,
            months=list(calendar.month_name)[1:]  # For currency converter
        )
        
    except Exception as e:
        flash('An error occurred while processing project', 'danger')
        logger.error(f"Error in manager_project_detail: {str(e)}", exc_info=True)
        return redirect(url_for('manager_projects'))
# Add these two routes for handling the ROI table and milestones views

@app.route('/manager/roi_table/<project_id>', methods=['GET', 'POST'])
@login_required(role='manager')
def manager_roi_table(project_id):
    try:
        project = db.projects.find_one({'project_id': project_id})
        if not project:
            flash('Project not found', 'danger')
            return redirect(url_for('manager_projects'))
        
        # Handle years_to_show instead of months_to_show
        try:
            years_to_show = int(request.form.get('years', 5)) if request.method == 'POST' else 5
            years_to_show = max(1, years_to_show)
        except ValueError:
            years_to_show = 5

        # Calculate ROI - first calculate monthly then convert to annual
        months_to_calculate = years_to_show * 12
        monthly_roi_data = calculate_roi({
            'project_type': project.get('project_type'),
            'project_parameters': project.get('project_parameters', []),
            'machinery_cost': project.get('machinery_cost')
        }, months_to_calculate)

        # Convert monthly data to annual
        annual_data = {
            'is_kaizen': monthly_roi_data.get('is_kaizen', False),
            'machinery_cost': monthly_roi_data.get('machinery_cost', 0),
            'annual_saving': monthly_roi_data.get('monthly_saving', 0) * 12,
            'roi_year': None,
            'years': []
        }

        # Group months into years
        for year_num in range(1, years_to_show + 1):
            start_month = (year_num - 1) * 12
            end_month = year_num * 12
            year_months = monthly_roi_data['months'][start_month:end_month]
            
            if not year_months:
                continue
                
            annual_saving = sum(month['saving'] for month in year_months)
            cumulative_saving = year_months[-1]['cumulative']
            is_roi = any(month['is_roi'] for month in year_months)
            
            if is_roi and annual_data['roi_year'] is None:
                annual_data['roi_year'] = year_num
                
            annual_data['years'].append({
                'year': year_num,
                'saving': annual_saving,
                'cumulative': cumulative_saving,
                'is_roi': is_roi
            })

        # Get currency data for converter
        current_year = datetime.now().year
        currencies = list(db.currency_rates.find(
            {'year': current_year}, 
            {'currency_name': 1, 'monthly_rates': 1}
        ).sort('currency_name', 1))

        # Get months for dropdown (still needed for currency conversion)
        months = list(calendar.month_name)[1:]  # ['January', 'February', ...]
        current_month = datetime.now().strftime('%B')

        return render_template(
            'manager/roi_table.html',
            project=project,
            roi_data=annual_data,
            years_to_show=years_to_show,
            currencies=currencies,
            months=months,
            current_month=current_month
        )
    except Exception as e:
        flash(f'An error occurred: {str(e)}', 'danger')
        return redirect(url_for('manager_projects'))
    
    
@app.route('/manager/milestones/<project_id>')
@login_required(role='manager')
def manager_milestones(project_id):
    try:
        # Get project data
        project = db.projects.find_one({'project_id': project_id})
        if not project:
            flash('Project not found', 'danger')
            return redirect(url_for('manager_projects'))
        
        # Get currency data for converter
        current_year = datetime.now().year
        currencies = list(db.currency_rates.find(
            {'year': current_year},
            {'currency_name': 1, 'monthly_rates': 1}
        ).sort('currency_name', 1))

        # Get months for dropdown
        months = list(calendar.month_name)[1:]  # ['January', 'February', ...]
        current_month = datetime.now().strftime('%B')

        # Get project parameters for cost savings calculation
        project_parameters = project.get('project_parameters', [])
        machinery_cost = project.get('machinery_cost', 0)

        # Process forecasted data
        forecasted_data = []
        forecasted_total = 0
        forecasted_cost_savings_total = 0
        for month_data in project.get('monthlyData', []):
            # Calculate total value - use monthly total if specified, otherwise sum daily values
            if month_data.get('userEnteredMonthly', False):
                total_value = month_data.get('totalValue', 0)
            else:
                total_value = sum(day.get('value', 0) for day in month_data.get('days', []))
            
            # Calculate cost savings for this month
            cost_savings = total_value * project_parameters[0]['cost'] if project_parameters else 0
            
            forecasted_total += total_value
            forecasted_cost_savings_total += cost_savings
            
            entry = {
                'month': month_data['month'],
                'year': month_data['year'],
                'totalValue': total_value,
                'costSavings': cost_savings,
                'userEnteredMonthly': month_data.get('userEnteredMonthly', False),
                'days': month_data.get('days', [])
            }
            forecasted_data.append(entry)
       
        # Process actual data
        actual_data = []
        actual_total = 0
        actual_cost_savings_total = 0
        for month_data in project.get('actualMonthlyData', []):
            if month_data.get('userEnteredMonthly', False):
                total_value = month_data.get('totalValue', 0)
            else:
                total_value = sum(day.get('value', 0) for day in month_data.get('days', []))
            
            # Calculate cost savings for this month
            cost_savings = total_value * project_parameters[0]['cost'] if project_parameters else 0
            
            actual_total += total_value
            actual_cost_savings_total += cost_savings
            
            entry = {
                'month': month_data['month'],
                'year': month_data['year'],
                'totalValue': total_value,
                'costSavings': cost_savings,
                'userEnteredMonthly': month_data.get('userEnteredMonthly', False),
                'days': month_data.get('days', [])
            }
            actual_data.append(entry)

        # Calculate ROI if we have both machinery cost and cost savings
        roi_forecast = None
        roi_actual = None
        if machinery_cost and forecasted_cost_savings_total:
            roi_forecast = ((forecasted_cost_savings_total - machinery_cost) / machinery_cost) * 100
        if machinery_cost and actual_cost_savings_total:
            roi_actual = ((actual_cost_savings_total - machinery_cost) / machinery_cost) * 100

        return render_template('manager/milestones.html',
                           project=project,
                           forecasted_data=forecasted_data,
                           actual_data=actual_data,
                           currencies=currencies,
                           months=months,
                           current_month=current_month,
                           project_parameters=project_parameters,
                           machinery_cost=machinery_cost,
                           forecasted_total=forecasted_total,
                           actual_total=actual_total,
                           forecasted_cost_savings_total=forecasted_cost_savings_total,
                           actual_cost_savings_total=actual_cost_savings_total,
                           roi_forecast=roi_forecast,
                           roi_actual=roi_actual)
    except Exception as e:
        flash(f'An error occurred: {str(e)}', 'danger')
        return redirect(url_for('manager_projects'))

def handle_project_updates(project_id, project):
    """Handle all POST request actions with validation"""
    action = request.form.get('action')
    comment = request.form.get('content', '').strip()
    
    # Validate action
    if action not in ['approve', 'reject', 'comment']:
        flash('Invalid action requested', 'danger')
        return redirect(url_for('manager_project_detail', project_id=project_id))
    
    # Validate comment for status changes
    if action in ['approve', 'reject'] and not comment:
        flash('Comment required for status changes', 'warning')
        return redirect(url_for('manager_project_detail', project_id=project_id))
    
    # Sanitize comment
    if len(comment) > 1000:
        comment = comment[:1000]
        flash('Comment was truncated to 1000 characters', 'warning')
    
    # Prepare update
    update_data = {
        'status': action.capitalize() if action in ['approve', 'reject'] else project['status'],
        'reviewed_at': datetime.now(),
        'reviewed_by': session.get('username', 'system')
    }
    
    new_comment = {
        'content': comment,
        'author': session.get('username', 'system'),
        'timestamp': datetime.now(),
        'type': action if action in ['approve', 'reject'] else 'comment'
    }
    
    try:
        result = db.projects.update_one(
            {'_id': ObjectId(project_id)},
            {'$set': update_data, '$push': {'comments': new_comment}}
        )
        
        if result.modified_count == 0:
            flash('No changes were made to the project', 'warning')
        else:
            flash(f'Project {action}d successfully!' if action in ['approve', 'reject'] else 'Comment added!', 'success')
            
    except Exception as e:
        logger.error(f"Failed to update project: {str(e)}")
        flash('Failed to update project', 'danger')
    
    return redirect(url_for('manager_project_detail', project_id=project_id))

def enhance_project_data(project):
    """Enrich project data with additional information"""
    # Convert ObjectId to string for template
    project['_id'] = str(project['_id'])
    
    # Add usernames with proper error handling
    for field in ['created_by', 'reviewed_by']:
        if field in project and isinstance(project[field], str):
            try:
                user = db.users.find_one({'username': project[field]}, {'name': 1})
                project[f'{field}_username'] = user.get('name', project[field]) if user else project[field]
            except Exception as e:
                logger.warning(f"Couldn't fetch user {project[field]}: {str(e)}")
                project[f'{field}_username'] = project[field]
    
    # Ensure lists exist
    for field in ['comments', 'project_parameters']:
        project.setdefault(field, [])
    
    # Convert and validate dates
    date_fields = ['start_date', 'end_date', 'created_at', 'reviewed_at']
    for date_field in date_fields:
        if date_field in project:
            try:
                if isinstance(project[date_field], str):
                    project[date_field] = datetime.strptime(project[date_field], '%Y-%m-%d')
                elif not isinstance(project[date_field], datetime):
                    project[date_field] = None
            except ValueError:
                project[date_field] = None
                logger.warning(f"Invalid date format in {date_field}")
                
    return project

def calculate_project_metrics(project):
    """Calculate financial metrics and timeline data with validation"""
    metrics = {
        'days_remaining': 0,
        'roi_percentage': 0,
        'total_savings': 0,
        'calculation_error': None
    }
    
    try:
        # Calculate savings and ROI
        machinery_cost = float(project.get('machinery_cost', 0)) or 1  # Avoid division by zero
        total_savings = 0.0
        
        for param in project.get('project_parameters', []):
            try:
                before = float(param.get('before_value', 0))
                after = float(param.get('after_value', 0))
                cost = float(param.get('cost', 0))
                
                savings = (before - after) * cost if param.get('operator') == '>' else (after - before) * cost
                total_savings += savings
            except (ValueError, TypeError):
                logger.warning(f"Invalid parameter values in project {project['_id']}")
                continue
        
        metrics['total_savings'] = round(total_savings, 2)
        metrics['roi_percentage'] = round((total_savings / machinery_cost * 100), 2) if machinery_cost > 0 else 0
        
        # Calculate days remaining
        if project.get('end_date') and isinstance(project['end_date'], datetime):
            metrics['days_remaining'] = max(0, (project['end_date'].date() - datetime.now().date()).days)
            
    except Exception as e:
        logger.error(f"Error calculating metrics for project {project.get('_id')}: {str(e)}")
        metrics['calculation_error'] = str(e)
    
    return metrics





import random

@app.route('/manager/dashboard')
@login_required(role='manager')
def manager_dashboard():
    # Your existing dashboard logic
    total_projects = db.projects.count_documents({})
    pending_reviews = db.projects.count_documents({'status': 'Pending'})
    approved_projects = db.projects.count_documents({'status': 'Approved'})
    
    all_projects = list(db.projects.find({'status': 'Approved'}))
    monthly_savings = sum(project.get('roi_data', {}).get('monthly_saving', 0) for project in all_projects)
    
    recent_projects = list(db.projects.find({'status': 'Pending'}).sort('created_at', -1).limit(5))
    
    return render_template('manager/manager_dashboard.html',
                         total_projects=total_projects,
                         pending_reviews=pending_reviews,
                         approved_projects=approved_projects,
                         monthly_savings=monthly_savings,
                         recent_projects=recent_projects)
from datetime import datetime
from collections import defaultdict
from flask import render_template

@app.route('/manager/analytics_dashboard')
@login_required(role='manager')
def manager_analytics_dashboard():
    """
    Analytics dashboard with mock data for demonstration
    """
    # Mock data for demonstration - ensure all values are simple data types (lists, dicts, numbers, strings)
    mock_data = {
        # Project Status Distribution
        'project_status_distribution': {
            'labels': ['Approved', 'Pending', 'In Progress', 'Rejected'],
            'values': [89, 34, 23, 10]
        },
        
        # Monthly Savings Trend (last 12 months)
        'monthly_savings_trend': {
            'labels': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'],
            'values': [120000, 135000, 148000, 162000, 175000, 189000, 
                      203000, 218000, 235000, 252000, 268000, 285000]
        },
        
        # Savings by Project Type
        'savings_by_type': {
            'labels': ['Six Sigma', 'Lean Manufacturing', 'Automation', 'Kaizen', 'Process Improvement'],
            'values': [450,000, 20,000, 280000, 180,000, 120,000]
        },
        
        # Forecast vs Actual Performance
        'forecast_vs_actual': {
           'labels': [
        'DMAIC Phase: Measure', 
        'DMAIC Phase: Analyze', 
        'DMAIC Phase: Improve', 
        'DMAIC Phase: Control', 
        'Lean Waste Reduction', 
        'Sigma Level Uplift'
    ],
    'actual': [145,000, 162,000, 175,000, 188,000, 205,000, 220,000],
    'forecasted': [150,000, 165,000, 180,000, 195,000, 210,000, 225,000]
        },
        
        # Six Sigma Phase Distribution
        'six_sigma_phases': {
            'labels': ['Define', 'Measure', 'Analyze', 'Improve', 'Control'],
            'values': [15, 12, 18, 22, 8]
        },
        
        # Six Sigma Projects Timeline
        'six_sigma_projects': {
            'labels': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'],
            'values': [8, 12, 15, 18, 22, 19, 16, 14, 11, 9, 7, 5]
        },
        
        # Key Metrics
        'total_projects': 156,
        'approved_projects': 89,
        'pending_projects': 34,
        'rejected_projects': 10,
        'in_progress_projects': 23,
        'total_budget': 2400000 if 2400000 > 0 else 1,  # $2.4M with fallback
        'total_machinery': 850000,  # $850K
        'total_savings': 1800000,  # $1.8M
        'active_projects': 123,
        'roi_percentage': 245,
        'payback_period_months': 18
    }
    
    try:
        return render_template('manager/analytics_dashboard.html', **mock_data)
    except Exception as e:
        print(f"Error in analytics dashboard: {str(e)}")
        # Return with safe default values that won't cause serialization errors
        safe_defaults = {
            'project_status_distribution': {'labels': [], 'values': []},
            'monthly_savings_trend': {'labels': [], 'values': []},
            'savings_by_type': {'labels': [], 'values': []},
            'forecast_vs_actual': {'labels': [], 'actual': [], 'forecasted': []},
            'six_sigma_phases': {'labels': [], 'values': []},
            'six_sigma_projects': {'labels': [], 'values': []},
            'total_projects': 1,
            'approved_projects': 1,
            'total_budget': 1,
            'total_machinery': 1,
            'total_savings': 1,
            'active_projects': 1,
            'roi_percentage': 0,
            'payback_period_months': 0
        }
        return render_template('manager/analytics_dashboard.html', **safe_defaults)
# PDF Report Generation

from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.legends import Legend
from reportlab.graphics.widgets.markers import makeMarker
from reportlab.lib.enums import TA_CENTER
from werkzeug.utils import secure_filename
import os
from datetime import datetime
import logging



def generate_report(project_id):
    try:
        project = db.projects.find_one({'project_id': project_id})
        if not project:
            raise Exception('Project not found')

        filename = secure_filename(f"Project_Report_{project.get('project_name', 'N/A')}.pdf")
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        # Use landscape orientation
        doc = SimpleDocTemplate(filepath, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        story = []

        # Custom styles
        title_style = ParagraphStyle('TitleStyle',
            parent=styles['Heading1'],
            alignment=TA_CENTER,
            fontSize=20,
            textColor=colors.maroon,
            fontName='Helvetica-Bold'
        )
        
        heading2_style = ParagraphStyle('Heading2Style',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.navy,
            fontName='Helvetica-Bold',
            spaceAfter=12
        )

        # Title
        story.append(Paragraph(f"Project Report: {project.get('project_name', 'N/A')}", title_style))
        story.append(Spacer(1, 12))

        # Project Details Section
        story.append(Paragraph("Project Details", heading2_style))
        project_details = [
            ("Project ID:", project.get('project_id', 'N/A')),
            ("Project Type:", project.get('project_type', 'N/A')),
            ("Factory:", project.get('factory_code', 'N/A')),
            ("Status:", project.get('status', 'N/A')),
            ("Created By:", project.get('created_by_username', 'N/A')),
            ("Created At:", project.get('created_at').strftime('%Y-%m-%d %H:%M') if isinstance(project.get('created_at'), datetime) else 'N/A'),
            ("Start Date:", project.get('start_date', 'N/A')),
            ("End Date:", project.get('end_date', 'N/A')),
            ("Description:", project.get('project_description', 'N/A'))
        ]
        
        project_table_data = [[Paragraph(label, styles['Normal']), Paragraph(str(value), styles['Normal'])] 
                            for label, value in project_details]
        project_table = Table(project_table_data, colWidths=[2 * inch, 4 * inch])
        project_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ]))
        story.append(project_table)
        story.append(Spacer(1, 24))

        # Cost Parameters Section
        story.append(Paragraph("Cost Parameters", heading2_style))
        project_parameters = project.get('project_parameters', [])
        if project_parameters:
            param_data = [["Parameter", "Before", "After", "Unit Cost"]]
            for param in project_parameters:
                param_data.append([
                    param.get('name', 'N/A'),
                    str(param.get('before_value', 'N/A')),
                    str(param.get('after_value', 'N/A')),
                    f"{param.get('cost', 'N/A')} {project.get('currency', 'PKR')}"
                ])
            
            param_table = Table(param_data, colWidths=[1.5*inch, 1*inch, 1*inch, 1.5*inch])
            param_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            story.append(param_table)
        else:
            story.append(Paragraph("No parameters defined.", styles['Normal']))
        story.append(Spacer(1, 24))

        # ROI Analysis Section
        story.append(Paragraph("ROI Analysis", heading2_style))
        roi_data = project.get('roi_data', {})
        
        # ROI Summary
        roi_summary = [
            ["Monthly Saving:", f"{roi_data.get('monthly_saving', 0):,.2f} {project.get('currency', 'PKR')}"],
            ["Machinery Cost:", f"{roi_data.get('machinery_cost', 0):,.2f} {project.get('currency', 'PKR')}"],
            ["Payback Period:", f"{roi_data.get('roi_month', 'N/A')} months"],
            ["Total Savings:", f"{roi_data.get('monthly_saving', 0) * 12:,.2f} {project.get('currency', 'PKR')}/year"]
        ]
        
        roi_summary_table = Table(roi_summary, colWidths=[2*inch, 3*inch])
        roi_summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ]))
        story.append(roi_summary_table)
        story.append(Spacer(1, 12))

        # ROI Monthly Table
        if 'months' in roi_data:
            story.append(Paragraph("Monthly ROI Projections", styles['Heading3']))
            roi_monthly_data = [["Month", "Monthly Saving", "Cumulative Saving"]]
            for month in roi_data['months']:
                roi_monthly_data.append([
                    month['month'],
                    f"{month['saving']:,.2f} {project.get('currency', 'PKR')}",
                    f"{month['cumulative']:,.2f} {project.get('currency', 'PKR')}"
                ])
            
            roi_monthly_table = Table(roi_monthly_data, colWidths=[1.5*inch, 2*inch, 2*inch])
            roi_monthly_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ]))
            story.append(roi_monthly_table)
        story.append(Spacer(1, 24))

        # Milestones Section with Charts
        story.append(PageBreak())
        story.append(Paragraph("Milestones Analysis", heading2_style))

        # Process data for charts
        forecasted_data = []
        actual_data = []
        labels = []
        
        for month_data in project.get('monthlyData', []):
            total_value = month_data.get('totalValue', 0) if month_data.get('userEnteredMonthly', False) else sum(day.get('value', 0) for day in month_data.get('days', []))
            forecasted_data.append(total_value)
            labels.append(month_data['month'][:3])  # Use first 3 letters of month name

        for month_data in project.get('actualMonthlyData', []):
            total_value = month_data.get('totalValue', 0) if month_data.get('userEnteredMonthly', False) else sum(day.get('value', 0) for day in month_data.get('days', []))
            actual_data.append(total_value)

        # Ensure both lists have same length
        max_length = max(len(forecasted_data), len(actual_data))
        forecasted_data += [0] * (max_length - len(forecasted_data))
        actual_data += [0] * (max_length - len(actual_data))

        # Create comparison chart (only if data exists)
        if forecasted_data or actual_data:
            story.append(Paragraph("Forecasted vs Actual Savings", styles['Heading3']))
            
            # Chart dimensions
            drawing = Drawing(400, 200)
            
            # Create bar chart
            bc = VerticalBarChart()
            bc.x = 50
            bc.y = 50
            bc.height = 150
            bc.width = 350
            bc.data = [forecasted_data or [0], actual_data or [0]]  # Handle empty data
            bc.strokeColor = colors.black
            
            # Safely calculate max value
            max_value = max(max(forecasted_data or [0]), max(actual_data or [0]))
            bc.valueAxis.valueMin = 0
            bc.valueAxis.valueMax = max_value * 1.2 if max_value > 0 else 100  # Fallback
            bc.valueAxis.valueStep = max(max_value / 5, 1)  # Ensure minimum step
            
            bc.categoryAxis.labels.boxAnchor = 'ne'
            bc.categoryAxis.labels.dx = 8
            bc.categoryAxis.labels.dy = -2
            bc.categoryAxis.labels.angle = 45
            bc.categoryAxis.categoryNames = labels or ["No Data"]  # Fallback
            
            bc.bars[0].fillColor = colors.blue
            bc.bars[1].fillColor = colors.green
            bc.bars[0].name = 'Forecasted'
            bc.bars[1].name = 'Actual'
            
            # Add legend
            legend = Legend()
            legend.alignment = 'right'
            legend.x = 350
            legend.y = 180
            legend.dxTextSpace = 5
            legend.columnMaximum = 2
            legend.colorNamePairs = [
                (colors.blue, 'Forecasted'),
                (colors.green, 'Actual')
            ]
            
            drawing.add(bc)
            drawing.add(legend)
            story.append(drawing)
            story.append(Spacer(1, 24))

            # Cumulative Savings Trend Chart
            story.append(Paragraph("Cumulative Savings Trend", styles['Heading3']))
            
            # Calculate cumulative data
            cum_forecast = []
            cum_actual = []
            total_f = 0
            total_a = 0
            
            for f, a in zip(forecasted_data, actual_data):
                total_f += f
                total_a += a
                cum_forecast.append(total_f)
                cum_actual.append(total_a)
            
            # Create line chart only if we have data
            if cum_forecast or cum_actual:
                drawing = Drawing(400, 200)
                lc = LinePlot()
                lc.x = 50
                lc.y = 50
                lc.height = 150
                lc.width = 350
                
                # Format data correctly for LinePlot: [(x1,y1), (x2,y2), ...]
                lc.data = [
                    list(zip(range(len(cum_forecast)), cum_forecast)),
                    list(zip(range(len(cum_actual)), cum_actual))
                ]
                
                lc.strokeColor = colors.black
                lc.lines[0].strokeColor = colors.blue
                lc.lines[1].strokeColor = colors.green
                lc.lines[0].strokeWidth = 2
                lc.lines[1].strokeWidth = 2
                lc.lines[0].symbol = makeMarker('Circle')
                lc.lines[1].symbol = makeMarker('Square')

                # Configure Y-axis
                lc.yValueAxis.valueMin = 0
                max_value = max(max(cum_forecast or [0]), max(cum_actual or [0]))
                lc.yValueAxis.valueMax = max_value * 1.2 if max_value > 0 else 100

                # Configure X-axis
                if labels:
                    lc.xValueAxis.labels.boxAnchor = 'ne'
                    lc.xValueAxis.labels.dx = 8
                    lc.xValueAxis.labels.dy = -2
                    lc.xValueAxis.labels.angle = 45
                    lc.xValueAxis.labels._text = labels  # Set labels directly
                
                # Add legend
                legend = Legend()
                legend.alignment = 'right'
                legend.x = 350
                legend.y = 180
                legend.dxTextSpace = 5
                legend.columnMaximum = 2
                legend.colorNamePairs = [
                    (colors.blue, 'Forecasted'),
                    (colors.green, 'Actual')
                ]
                
                drawing.add(lc)
                drawing.add(legend)
                story.append(drawing)
            else:
                story.append(Paragraph("No cumulative data available for trend analysis.", styles['Normal']))

            story.append(Spacer(1, 24))

            # Milestones Comparison Table
            story.append(Paragraph("Savings Comparison", styles['Heading3']))
            milestones_data = [["Month", "Forecasted", "Actual", "Variance"]]
            
            for i in range(max(len(forecasted_data), len(actual_data))):
                forecast = forecasted_data[i] if i < len(forecasted_data) else None
                actual = actual_data[i] if i < len(actual_data) else None
                
                month = labels[i] if i < len(labels) else f"Month {i+1}"
                forecast_val = forecast if forecast else 0
                actual_val = actual if actual else 0
                variance = actual_val - forecast_val
                
                milestones_data.append([
                    month,
                    f"{forecast_val:,.2f} {project.get('currency', 'PKR')}" if forecast else "N/A",
                    f"{actual_val:,.2f} {project.get('currency', 'PKR')}" if actual else "N/A",
                    f"{variance:,.2f} {project.get('currency', 'PKR')}"
                ])
            
            milestones_table = Table(milestones_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            milestones_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (-1, 1), (-1, -1), lambda r, c, data=len(milestones_data): colors.green if r > 0 and float(milestones_data[r][3].split()[0]) >= 0 else colors.red),
            ]))
            story.append(milestones_table)
            story.append(Spacer(1, 12))
        else:
            story.append(Paragraph("No forecasted or actual data available.", styles['Normal']))
            story.append(Spacer(1, 24))

        # Detailed Daily Data (for first 3 months or all if less than 3)
        months_to_show = min(3, len(project.get('actualMonthlyData', [])))
        for i in range(months_to_show):
            month_data = project['actualMonthlyData'][i]
            if not month_data.get('userEnteredMonthly', True):  # Only show daily breakdown if it exists
                story.append(PageBreak())
                story.append(Paragraph(f"Daily Savings - {month_data['month']} {month_data['year']}", heading2_style))
                
                daily_data = [["Day", "Savings", "Notes"]]
                for day in month_data.get('days', []):
                    daily_data.append([
                        str(day['day']),
                        f"{day['value']:,.2f} {project.get('currency', 'PKR')}",
                        day.get('notes', '')[:50]  # Limit notes length
                    ])
                
                daily_table = Table(daily_data, colWidths=[1*inch, 2*inch, 3*inch])
                daily_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                ]))
                story.append(daily_table)
                story.append(Spacer(1, 12))

        # Approval Details (if approved)
        if project.get('status') == 'Approved':
            story.append(PageBreak())
            story.append(Paragraph("Approval Details", heading2_style))
            approval_details = [
                ["Reviewed By:", project.get('reviewed_by_username', 'N/A')],
                ["Reviewed At:", project.get('reviewed_at').strftime('%Y-%m-%d %H:%M') 
                 if isinstance(project.get('reviewed_at'), datetime) else 'N/A'],
                ["Comments:", project.get('manager_comments', 'No comments provided.')]
            ]
            approval_table = Table(approval_details, colWidths=[2*inch, 4*inch])
            approval_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(approval_table)

        doc.build(story)
        return filepath

    except Exception as e:
        logger.error(f"Error generating report: {str(e)}")
        raise Exception(f"Failed to generate report: {str(e)}")
    
@app.route('/download-example-csv')
@login_required(role='admin')
def download_example_csv():
    try:
        exports_dir = os.path.join(app.root_path, 'static', 'exports')
        filename = 'example_factory_data.csv'  # Same file or different if you have another
        
        if not os.path.exists(os.path.join(exports_dir, filename)):
            flash("Example CSV file not found. Please contact support.", "danger")
            return redirect(url_for('admin_upload'))
        
        return send_from_directory(
            directory=exports_dir,
            path=filename,
            as_attachment=True,
            download_name='example_factory_data.csv',
            mimetype='text/csv'
        )
    except Exception as e:
        app.logger.error(f"Error downloading example CSV: {str(e)}")
        flash("Could not download example file. Please contact support.", "danger")
        return redirect(url_for('admin_upload'))


@app.route('/user/delete_project/<project_id>', methods=['POST'])
@login_required(role='user')
def user_delete_project(project_id):
    try:
        result = db.projects.delete_one({'project_id': project_id, 'created_by_user_id': session['user_id']})
        if result.deleted_count > 0:
            flash('Project deleted successfully!', 'success')
        else:
            flash('Project not found or you do not have permission to delete it.', 'danger')
    except Exception as e:
        flash(f'Error deleting project: {e}', 'danger')
    return redirect(url_for('user_dashboard'))

if __name__ == '__main__':
    with app.app_context():
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        initialize_database()

    app.run(debug=True)